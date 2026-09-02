"""Вкладка "Документи" (PySide6) — генерація Excel і зберігання в PostgreSQL.

ВИПРАВЛЕННЯ: документи зберігаються в БД (project_documents), а не у файловій системі.
"""

from PySide6.QtCore import Qt
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QComboBox, QMessageBox, QGridLayout
)

from ventilation_company.gui_pyside6.theme import Theme
from ventilation_company.database.db import get_db
from ventilation_company.database.models.project import Project
from ventilation_company.database.repositories.product_repo import ProductRepository
from ventilation_company.database.repositories.project_document_repo import ProjectDocumentRepository

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime
import io


class DocumentsTab(QWidget):
    """Вкладка генерації документів зі зберіганням у БД."""

    def __init__(self, parent=None, main_window=None):
        super().__init__(parent)
        self.main_window = main_window
        self._current_project_id = None
        self._build_ui()
        self._load_projects()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(24, 24, 24, 24)
        layout.setSpacing(16)

        top = QHBoxLayout()
        lbl_title = QLabel("📄 Документи")
        lbl_title.setObjectName("title")
        top.addWidget(lbl_title)
        top.addSpacing(24)

        lbl_proj = QLabel("Проєкт:")
        lbl_proj.setStyleSheet(f"color: {Theme.TEXT_MUTED};")
        top.addWidget(lbl_proj)

        self.combo_project = QComboBox()
        self.combo_project.setMinimumWidth(300)
        self.combo_project.setMinimumHeight(32)
        self.combo_project.currentIndexChanged.connect(self._on_project_changed)
        top.addWidget(self.combo_project)
        top.addStretch()
        layout.addLayout(top)

        lbl_docs = QLabel("Оберіть тип документа для генерації:")
        lbl_docs.setStyleSheet(f"color: {Theme.TEXT_BRIGHT}; font-size: 14px;")
        layout.addWidget(lbl_docs)

        docs_grid = QGridLayout()
        docs_grid.setSpacing(12)

        self.btn_spec = QPushButton("📋 Специфікація\nдля замовника")
        self.btn_spec.setMinimumHeight(80)
        self.btn_spec.setStyleSheet(f"font-size: 13px; background: {Theme.BG_CARD}; border: 1px solid {Theme.BORDER}; border-radius: 8px;")
        self.btn_spec.clicked.connect(lambda: self._generate("spec"))
        docs_grid.addWidget(self.btn_spec, 0, 0)

        self.btn_calc = QPushButton("🧮 Калькуляція\nсобівартості")
        self.btn_calc.setMinimumHeight(80)
        self.btn_calc.setStyleSheet(f"font-size: 13px; background: {Theme.BG_CARD}; border: 1px solid {Theme.BORDER}; border-radius: 8px;")
        self.btn_calc.clicked.connect(lambda: self._generate("calc"))
        docs_grid.addWidget(self.btn_calc, 0, 1)

        self.btn_metal = QPushButton("🔩 Замовлення\nна метал")
        self.btn_metal.setMinimumHeight(80)
        self.btn_metal.setStyleSheet(f"font-size: 13px; background: {Theme.BG_CARD}; border: 1px solid {Theme.BORDER}; border-radius: 8px;")
        self.btn_metal.clicked.connect(lambda: self._generate("metal"))
        docs_grid.addWidget(self.btn_metal, 1, 0)

        self.btn_order = QPushButton("🏭 Наряд на\nвиробництво")
        self.btn_order.setMinimumHeight(80)
        self.btn_order.setStyleSheet(f"font-size: 13px; background: {Theme.BG_CARD}; border: 1px solid {Theme.BORDER}; border-radius: 8px;")
        self.btn_order.clicked.connect(lambda: self._generate("order"))
        docs_grid.addWidget(self.btn_order, 1, 1)

        layout.addLayout(docs_grid)

        info = QLabel("💡 Документи зберігаються в базі даних PostgreSQL\n"
                      "Перегляньте їх у картці проєкту (вкладка 'Проєкти' → двічі клікніть на проєкт)")
        info.setStyleSheet(f"color: {Theme.TEXT_MUTED}; font-size: 11px; padding: 8px;")
        info.setWordWrap(True)
        layout.addWidget(info)
        layout.addStretch()

    def _load_projects(self):
        self.combo_project.blockSignals(True)
        self.combo_project.clear()
        try:
            with get_db() as session:
                projects = session.query(Project).order_by(Project.created_at.desc()).all()
                for p in projects:
                    display = f"{p.project_number or '—'} — {p.name or 'Без назви'}"
                    self.combo_project.addItem(display, p.id)
        except Exception as e:
            QMessageBox.critical(self, "Помилка", f"Не вдалося завантажити проєкти: {e}")
        self.combo_project.blockSignals(False)

        active_id = self.main_window.active_project_id if self.main_window else None
        if active_id:
            idx = self.combo_project.findData(active_id)
            if idx >= 0:
                self.combo_project.setCurrentIndex(idx)
                return
        if self.combo_project.count() > 0:
            self.combo_project.setCurrentIndex(0)
            self._current_project_id = self.combo_project.itemData(0)

    def _on_project_changed(self, index):
        self._current_project_id = self.combo_project.itemData(index)

    def _get_project_data(self):
        if not self._current_project_id:
            QMessageBox.warning(self, "Увага", "Спочатку виберіть проєкт")
            return None, None
        try:
            with get_db() as session:
                p = session.query(Project).filter(Project.id == self._current_project_id).first()
                if not p:
                    QMessageBox.warning(self, "Увага", "Проєкт не знайдено")
                    return None, None
                project_data = {
                    "id": p.id,
                    "name": p.name or "—",
                    "project_number": p.project_number or str(p.id),
                    "client": p.client or "—",
                }
                products = ProductRepository.get_all(project_id=self._current_project_id)
                return project_data, products
        except Exception as e:
            QMessageBox.critical(self, "Помилка", f"Не вдалося завантажити дані: {e}")
            return None, None

    def _generate(self, doc_type: str):
        project, products = self._get_project_data()
        if not project:
            return
        if not products:
            QMessageBox.warning(self, "Увага", "У проєкті немає виробів.")
            return

        try:
            # Генеруємо Excel у пам'ять
            buffer = io.BytesIO()
            filename = f"{doc_type}_{project['project_number']}_{datetime.now().strftime('%Y%m%d')}.xlsx"

            if doc_type == "spec":
                self._gen_spec(project, products, buffer)
            elif doc_type == "calc":
                self._gen_calc(project, products, buffer)
            elif doc_type == "metal":
                self._gen_metal(project, products, buffer)
            elif doc_type == "order":
                self._gen_order(project, products, buffer)

            content = buffer.getvalue()

            # Зберігаємо в БД
            ProjectDocumentRepository.create(
                project_id=project["id"],
                doc_type=doc_type,
                filename=filename,
                content=content,
            )

            QMessageBox.information(self, "Успіх",
                f"Документ збережено в базі даних!\n\n"
                f"Файл: {filename}\n"
                f"Розмір: {len(content) / 1024:.1f} КБ\n\n"
                f"Перегляньте у картці проєкту (вкладка 'Проєкти').")
        except Exception as e:
            QMessageBox.critical(self, "Помилка", f"Не вдалося згенерувати документ: {e}")

    def _style_header(self, ws, row, headers, fill_color="4472C4"):
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        font = Font(bold=True, color="FFFFFF")
        border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

    def _gen_spec(self, project, products, buffer):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Специфікація"
        ws.merge_cells("A1:H1")
        ws["A1"] = f"СПЕЦИФІКАЦІЯ № {project['project_number']}"
        ws["A1"].font = Font(bold=True, size=16)
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.merge_cells("A2:H2")
        ws["A2"] = f"Проєкт: {project['name']} | Клієнт: {project['client']}"
        ws["A2"].alignment = Alignment(horizontal="center")
        headers = ["№", "Назва", "Тип", "Розміри", "Матеріал", "К-ть", "Ціна", "Сума"]
        self._style_header(ws, 4, headers)
        total = 0
        for i, item in enumerate(products, 1):
            w = item.get("width", 0) or 0
            h = item.get("height", 0) or 0
            l = item.get("length", 0) or 0
            dims = f"Ø{w:.0f} x {l:.0f}" if h == 0 else f"{w:.0f}x{h:.0f}x{l:.0f}"
            row_data = [i, item.get("name"), item.get("product_type"), dims, item.get("material"), item.get("quantity", 1), item.get("unit_price", 0), item.get("total_price", 0)]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=4+i, column=col, value=val)
                cell.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
            total += item.get("total_price", 0)
        row = 4 + len(products) + 1
        ws.merge_cells(f"A{row}:G{row}")
        ws.cell(row=row, column=1, value="ВСЬОГО:").font = Font(bold=True)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="right")
        ws.cell(row=row, column=8, value=total).font = Font(bold=True)
        for col in range(1, 9):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["D"].width = 18
        wb.save(buffer)

    def _gen_calc(self, project, products, buffer):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Калькуляція"
        ws.merge_cells("A1:F1")
        ws["A1"] = f"КАЛЬКУЛЯЦІЯ СОБІВАРТОСТІ — {project['name']}"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")
        headers = ["№", "Назва", "Матеріал", "К-ть", "Собівартість", "Кінцева ціна"]
        self._style_header(ws, 3, headers, "70AD47")
        total_cost = 0
        total_price = 0
        for i, item in enumerate(products, 1):
            cost = item.get("unit_price", 0) * 0.7
            price = item.get("total_price", 0)
            row_data = [i, item.get("name"), item.get("material"), item.get("quantity", 1), round(cost, 2), price]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=3+i, column=col, value=val)
                cell.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
            total_cost += cost
            total_price += price
        row = 3 + len(products) + 1
        ws.merge_cells(f"A{row}:D{row}")
        ws.cell(row=row, column=1, value="ВСЬОГО:").font = Font(bold=True)
        ws.cell(row=row, column=5, value=round(total_cost, 2)).font = Font(bold=True)
        ws.cell(row=row, column=6, value=total_price).font = Font(bold=True)
        profit = total_price - total_cost
        row += 1
        ws.merge_cells(f"A{row}:D{row}")
        ws.cell(row=row, column=1, value="ПРИБУТОК:").font = Font(bold=True, color="006100")
        ws.cell(row=row, column=5, value=round(profit, 2)).font = Font(bold=True, color="006100")
        wb.save(buffer)

    def _gen_metal(self, project, products, buffer):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Метал"
        ws.merge_cells("A1:E1")
        ws["A1"] = f"ЗАМОВЛЕННЯ НА МЕТАЛ — {project['name']}"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")
        metal_summary = {}
        for item in products:
            mat = item.get("material", "—")
            thick = item.get("thickness", "—")
            area = item.get("metal_area_m2", 0) or 0
            weight = item.get("weight_kg", 0) or 0
            key = f"{mat} | {thick} мм"
            if key not in metal_summary:
                metal_summary[key] = {"area": 0, "weight": 0, "qty": 0}
            metal_summary[key]["area"] += area * item.get("quantity", 1)
            metal_summary[key]["weight"] += weight * item.get("quantity", 1)
            metal_summary[key]["qty"] += item.get("quantity", 1)
        headers = ["№", "Матеріал | Товщина", "Площа м²", "Вага кг", "К-ть виробів"]
        self._style_header(ws, 3, headers, "FFC000")
        total_area = 0
        total_weight = 0
        for i, (key, data) in enumerate(metal_summary.items(), 1):
            row_data = [i, key, round(data["area"], 2), round(data["weight"], 2), data["qty"]]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=3+i, column=col, value=val)
                cell.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
            total_area += data["area"]
            total_weight += data["weight"]
        row = 3 + len(metal_summary) + 1
        ws.merge_cells(f"A{row}:B{row}")
        ws.cell(row=row, column=1, value="ВСЬОГО:").font = Font(bold=True)
        ws.cell(row=row, column=3, value=round(total_area, 2)).font = Font(bold=True)
        ws.cell(row=row, column=4, value=round(total_weight, 2)).font = Font(bold=True)
        wb.save(buffer)

    def _gen_order(self, project, products, buffer):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Наряд"
        ws.merge_cells("A1:G1")
        ws["A1"] = f"НАРЯД НА ВИРОБНИЦТВО — {project['name']}"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")
        headers = ["№", "Назва", "Тип", "Розміри", "Матеріал", "Товщ.", "К-ть"]
        self._style_header(ws, 3, headers, "C65911")
        for i, item in enumerate(products, 1):
            w = item.get("width", 0) or 0
            h = item.get("height", 0) or 0
            l = item.get("length", 0) or 0
            dims = f"Ø{w:.0f} x {l:.0f}" if h == 0 else f"{w:.0f}x{h:.0f}x{l:.0f}"
            row_data = [i, item.get("name"), item.get("product_type"), dims, item.get("material"), item.get("thickness", "—"), item.get("quantity", 1)]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=3+i, column=col, value=val)
                cell.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        wb.save(buffer)

    def refresh(self):
        self._load_projects()

    def on_project_changed(self, project_id):
        if project_id:
            idx = self.combo_project.findData(project_id)
            if idx >= 0:
                self.combo_project.setCurrentIndex(idx)
        else:
            self.combo_project.setCurrentIndex(-1)
            self._current_project_id = None
