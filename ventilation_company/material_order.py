"""Розрахунок потреби в матеріалах та генерація заявки для постачальника.

Підтримує:
  • Листовий метал (по товщині, матеріалу, розміру листа)
  • Ущільнювачі (профіль, стрічка)
  • Болти/гайки/шайби (для фланців)
  • Ізоляція (мінвата, каучук)
  • Комплектуючі (вентилятори, фільтри, клапани тощо)
  • Гнучкі вставки, решітки, дифузори

Генерація Excel-файлу заявки через openpyxl.
"""

import os
from dataclasses import dataclass, field
from datetime import datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter

from ventilation_company.config import MATERIALS, COMPONENTS


@dataclass
class MaterialItem:
    """Один рядок заявки на матеріали."""
    category: str           # "Листовий метал", "Ущільнювачі", "Кріплення", "Ізоляція", "Комплектуючі"
    name: str             # Назва матеріалу
    specification: str    # Специфікація (товщина, розмір, тип)
    unit: str             # Одиниці виміру
    quantity: float       # Кількість
    price_per_unit: float = 0.0   # Ціна за одиницю (опціонально)
    supplier: str = ""   # Постачальник
    notes: str = ""       # Примітки

    @property
    def total_price(self) -> float:
        return self.quantity * self.price_per_unit


@dataclass
class MaterialOrder:
    """Повна заявка на матеріали."""
    project_name: str
    order_date: datetime = field(default_factory=datetime.now)
    items: list[MaterialItem] = field(default_factory=list)
    notes: str = ""

    @property
    def total_items(self) -> int:
        return len(self.items)

    @property
    def total_cost(self) -> float:
        return sum(i.total_price for i in self.items)

    def get_by_category(self, category: str) -> list[MaterialItem]:
        return [i for i in self.items if i.category == category]


class MaterialCalculator:
    """Розрахунок потреби в матеріалах зі списку виробів."""

    # ── КОНСТАНТИ РОЗХОДУ МАТЕРІАЛІВ ──

    # Ущільнювач для фланця: м/п на один фланець
    GASKET_PER_FLANGE_M = 1.2

    # Болтів на один фланець (залежить від розміру)
    BOLTS_SMALL = 4    # для фланців < 400 мм
    BOLTS_MEDIUM = 6   # для фланців 400-800 мм
    BOLTS_LARGE = 8    # для фланців > 800 мм

    # Ізоляція: коефіцієнт на втрати при монтажі
    INSULATION_WASTE_FACTOR = 1.15

    # Стандартні розміри листів
    SHEET_SIZES = {
        (1250, 2500): "1250×2500",
        (1000, 2000): "1000×2000",
        (1500, 3000): "1500×3000",
    }

    def __init__(self):
        self.items: list[MaterialItem] = []

    def _get_sheet_size(self, thickness: float, material: str) -> str:
        """Визначити оптимальний розмір листа."""
        # За замовчуванням 1250×2500 для всіх
        return "1250×2500"

    def _get_bolt_count(self, width: float, height: float) -> int:
        """Кількість болтів на фланець залежно від розміру."""
        max_dim = max(width, height)
        if max_dim < 400:
            return self.BOLTS_SMALL
        elif max_dim <= 800:
            return self.BOLTS_MEDIUM
        else:
            return self.BOLTS_LARGE

    def calculate(self, products: list[dict]) -> MaterialOrder:
        """Розрахувати потребу в матеріалах для списку виробів."""
        self.items = []

        # Агрегація по металу
        metal_summary = {}   # (material, thickness) -> (total_area_m2, count)
        flange_summary = {}  # (width, height) -> count
        insulation_m2 = 0.0
        components_needed = {}  # component_name -> quantity

        for p in products:
            qty = int(p.get("quantity", 1))
            ptype = p.get("product_type", "").lower()
            material_str = p.get("material", "оцинкована сталь")
            thickness = float(p.get("thickness", 0.7))
            width = float(p.get("width", 0))
            height = float(p.get("height", 0))
            length = float(p.get("length", 0))
            area = float(p.get("metal_area_m2", 0))

            # ── Листовий метал ──
            if area > 0:
                key = (material_str, thickness)
                if key not in metal_summary:
                    metal_summary[key] = [0.0, 0]
                metal_summary[key][0] += area * qty
                metal_summary[key][1] += qty

            # ── Фланці та кріплення ──
            has_flanges = p.get("has_flanges", False)
            flange_count = int(p.get("flange_count", 0))
            if has_flanges and flange_count > 0:
                total_flanges = flange_count * qty
                fkey = (width, height)
                if fkey not in flange_summary:
                    flange_summary[fkey] = 0
                flange_summary[fkey] += total_flanges

            # ── Ізоляція (тільки для повітропроводів) ──
            if "duct" in ptype or "повітропровід" in ptype.lower():
                duct_area = 2 * (width + height) / 1000 * length / 1000  # м² зовнішньої поверхні
                insulation_m2 += duct_area * qty

            # ── Комплектуючі ──
            # Якщо виріб містить комплектуючі
            comp = p.get("components", [])
            if isinstance(comp, list):
                for c in comp:
                    cname = c if isinstance(c, str) else c.get("name", "")
                    if cname:
                        components_needed[cname] = components_needed.get(cname, 0) + qty

        # ── Формуємо рядки заявки ──

        # 1. Листовий метал
        for (material, thickness), (area, count) in metal_summary.items():
            sheet_size = self._get_sheet_size(thickness, material)
            # Розрахуємо кількість листів (з запасом 10%)
            sheet_area_m2 = 1.25 * 2.5  # 3.125 м² для 1250×2500
            sheets_needed = int((area * 1.10 / sheet_area_m2) + 0.999)  # округлення вгору

            material_key = f"{material.replace(' ', '_')}_{thickness}"
            price = MATERIALS.get(material_key, {}).get("ціна_за_м2", 0)

            self.items.append(MaterialItem(
                category="Листовий метал",
                name=f"{material} {thickness} мм",
                specification=f"Лист {sheet_size} мм",
                unit="шт",
                quantity=sheets_needed,
                price_per_unit=round(price * sheet_area_m2, 2) if price else 0,
                notes=f"Площа виробів: {area:.2f} м², виробів: {count} шт",
            ))

        # 2. Ущільнювачі для фланців
        total_gasket_m = 0
        total_bolts = 0
        total_nuts = 0
        total_washers = 0

        for (width, height), count in flange_summary.items():
            gasket_m = self.GASKET_PER_FLANGE_M * count
            total_gasket_m += gasket_m

            bolt_count = self._get_bolt_count(width, height) * count
            total_bolts += bolt_count
            total_nuts += bolt_count
            total_washers += bolt_count * 2  # 2 шайби на болт

        if total_gasket_m > 0:
            self.items.append(MaterialItem(
                category="Ущільнювачі",
                name="Ущільнювальний профіль EPDM",
                specification="Профіль 10×15 мм для фланців",
                unit="м.п.",
                quantity=round(total_gasket_m, 1),
                price_per_unit=25.0,
                notes=f"Фланців: {sum(flange_summary.values())} шт",
            ))

        # 3. Кріплення
        if total_bolts > 0:
            self.items.append(MaterialItem(
                category="Кріплення",
                name="Болт з шестигранною головкою",
                specification="М8×30 мм, оцинкований, DIN 933",
                unit="шт",
                quantity=total_bolts,
                price_per_unit=3.5,
            ))
            self.items.append(MaterialItem(
                category="Кріплення",
                name="Гайка шестигранна",
                specification="М8, оцинкована, DIN 934",
                unit="шт",
                quantity=total_nuts,
                price_per_unit=1.2,
            ))
            self.items.append(MaterialItem(
                category="Кріплення",
                name="Шайба плоска",
                specification="М8, оцинкована, DIN 125",
                unit="шт",
                quantity=total_washers,
                price_per_unit=0.5,
            ))

        # 4. Ізоляція
        if insulation_m2 > 0:
            insulated_m2 = insulation_m2 * self.INSULATION_WASTE_FACTOR
            self.items.append(MaterialItem(
                category="Ізоляція",
                name="Мінеральна вата",
                specification="ISOVER Венті 50 мм, 1000×600 мм",
                unit="м²",
                quantity=round(insulated_m2, 1),
                price_per_unit=MATERIALS.get("ізоляція_мінвата", {}).get("ціна_за_м2", 180),
                notes="З урахуванням 15% відходів",
            ))
            self.items.append(MaterialItem(
                category="Ізоляція",
                name="Склотканина",
                specification="Алюмінізована, 50 мм",
                unit="м²",
                quantity=round(insulated_m2, 1),
                price_per_unit=45.0,
                notes="Для обгортки ізоляції",
            ))

        # 5. Комплектуючі
        for comp_name, qty in components_needed.items():
            comp_key = comp_name.lower().replace(" ", "_")
            price = COMPONENTS.get(comp_key, {}).get("ціна", 0)
            unit = COMPONENTS.get(comp_key, {}).get("одиниця", "шт")
            self.items.append(MaterialItem(
                category="Комплектуючі",
                name=comp_name,
                specification="Згідно специфікації проєкту",
                unit=unit,
                quantity=qty,
                price_per_unit=price,
            ))

        # 6. Загальні матеріали (завжди додаємо)
        self.items.append(MaterialItem(
            category="Розхідні матеріали",
            name="Електроди зварювальні",
            specification="ОЗЛ-6, d=3 мм",
            unit="кг",
            quantity=5.0,
            price_per_unit=85.0,
            notes="Приблизна потреба",
        ))
        self.items.append(MaterialItem(
            category="Розхідні матеріали",
            name="Фарба алкідна",
            specification="ПФ-115, сіра, 2.5 кг",
            unit="шт",
            quantity=2,
            price_per_unit=320.0,
            notes="Для фарбування виробів",
        ))
        self.items.append(MaterialItem(
            category="Розхідні матеріали",
            name="Герметик силіконовий",
            specification="Нейтральний, 300 мл, прозорий",
            unit="шт",
            quantity=5,
            price_per_unit=65.0,
            notes="Для ущільнення з'єднань",
        ))

        return MaterialOrder(
            project_name="Заявка на матеріали",
            items=self.items,
        )


class MaterialOrderExporter:
    """Експорт заявки на матеріали в Excel."""

    def __init__(self, order: MaterialOrder):
        self.order = order

    def _setup_styles(self, wb: Workbook):
        """Налаштувати стилі для Excel."""
        # Заголовок
        header_style = NamedStyle(name="header")
        header_style.font = Font(bold=True, size=12, color="FFFFFF")
        header_style.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        header_style.alignment = Alignment(horizontal="center", vertical="center")
        wb.add_named_style(header_style)

        # Підзаголовок категорії
        cat_style = NamedStyle(name="category")
        cat_style.font = Font(bold=True, size=11, color="FFFFFF")
        cat_style.fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
        cat_style.alignment = Alignment(horizontal="left", vertical="center")
        wb.add_named_style(cat_style)

        # Заголовок таблиці
        th_style = NamedStyle(name="table_header")
        th_style.font = Font(bold=True, size=10, color="FFFFFF")
        th_style.fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
        th_style.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        wb.add_named_style(th_style)

        # Звичайна комірка
        cell_style = NamedStyle(name="cell")
        cell_style.font = Font(size=10)
        cell_style.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        wb.add_named_style(cell_style)

        # Числова комірка
        num_style = NamedStyle(name="number")
        num_style.font = Font(size=10)
        num_style.alignment = Alignment(horizontal="right", vertical="center")
        num_style.number_format = "#,##0.00"
        wb.add_named_style(num_style)

        # Ціна
        price_style = NamedStyle(name="price")
        price_style.font = Font(size=10)
        price_style.alignment = Alignment(horizontal="right", vertical="center")
        price_style.number_format = "#,##0.00"
        wb.add_named_style(price_style)

        # Ітого
        total_style = NamedStyle(name="total")
        total_style.font = Font(bold=True, size=11)
        total_style.fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
        total_style.alignment = Alignment(horizontal="right", vertical="center")
        total_style.number_format = "#,##0.00"
        wb.add_named_style(total_style)

    def export(self, filepath: str) -> str:
        """Зберегти заявку в Excel-файл."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Заявка на матеріали"

        self._setup_styles(wb)

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # ── ЗАГОЛОВОК ──
        ws.merge_cells("A1:H1")
        ws["A1"] = f"ЗАЯВКА НА МАТЕРІАЛИ — {self.order.project_name}"
        ws["A1"].style = "header"
        ws.row_dimensions[1].height = 30

        ws.merge_cells("A2:H2")
        ws["A2"] = f"Дата формування: {self.order.order_date.strftime('%d.%m.%Y %H:%M')}"
        ws["A2"].font = Font(italic=True, size=10)
        ws["A2"].alignment = Alignment(horizontal="center")

        row = 4

        # ── ЗАГОЛОВКИ КОЛОНОК ──
        headers = ["№", "Категорія", "Найменування", "Специфікація", "Од. вим.", "Кількість", "Ціна", "Сума"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.style = "table_header"
            cell.border = thin_border
        ws.row_dimensions[row].height = 25

        # ── ДАНІ ──
        categories = {}
        for item in self.order.items:
            cat = item.category
            if cat not in categories:
                categories[cat] = []
            categories[cat].append(item)

        row_num = 1
        grand_total = 0.0

        for cat_name in sorted(categories.keys()):
            items = categories[cat_name]
            cat_total = 0.0

            # Рядок категорії
            row += 1
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
            ws.cell(row=row, column=2, value=cat_name.upper()).style = "category"
            for col in range(1, 9):
                ws.cell(row=row, column=col).border = thin_border
            ws.row_dimensions[row].height = 22

            for item in items:
                row += 1
                ws.cell(row=row, column=1, value=row_num).style = "cell"
                ws.cell(row=row, column=2, value=item.category).style = "cell"
                ws.cell(row=row, column=3, value=item.name).style = "cell"
                ws.cell(row=row, column=4, value=item.specification).style = "cell"
                ws.cell(row=row, column=5, value=item.unit).style = "cell"
                ws.cell(row=row, column=5).alignment = Alignment(horizontal="center")
                ws.cell(row=row, column=6, value=item.quantity).style = "number"
                ws.cell(row=row, column=7, value=item.price_per_unit).style = "price"
                ws.cell(row=row, column=8, value=item.total_price).style = "price"

                for col in range(1, 9):
                    ws.cell(row=row, column=col).border = thin_border

                row_num += 1
                cat_total += item.total_price

            # Підсумок по категорії
            row += 1
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
            ws.cell(row=row, column=1, value=f"Разом по категорії «{cat_name}»:").style = "total"
            ws.cell(row=row, column=1).alignment = Alignment(horizontal="right")
            ws.cell(row=row, column=8, value=cat_total).style = "total"
            for col in range(1, 9):
                ws.cell(row=row, column=col).border = thin_border
            ws.row_dimensions[row].height = 22

            grand_total += cat_total
            row += 1  # порожній рядок

        # ── ЗАГАЛЬНИЙ ПІДСУМОК ──
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        ws.cell(row=row, column=1, value="ЗАГАЛЬНА СУМА:").style = "total"
        ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="C0392B")
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="right")
        ws.cell(row=row, column=8, value=grand_total).style = "total"
        ws.cell(row=row, column=8).font = Font(bold=True, size=12, color="C0392B")
        for col in range(1, 9):
            ws.cell(row=row, column=col).border = thin_border
        ws.row_dimensions[row].height = 28

        # ── ПРИМІТКИ ──
        if self.order.notes:
            row += 2
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
            ws.cell(row=row, column=1, value=f"Примітки: {self.order.notes}")
            ws.cell(row=row, column=1).font = Font(italic=True, size=9)
            ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)

        # ── ПІДПИСИ ──
        row += 3
        ws.cell(row=row, column=1, value="Склав:").font = Font(size=10)
        ws.cell(row=row, column=3, value="_________________").font = Font(size=10)
        row += 1
        ws.cell(row=row, column=1, value="Перевірив:").font = Font(size=10)
        ws.cell(row=row, column=3, value="_________________").font = Font(size=10)
        row += 1
        ws.cell(row=row, column=1, value="Директор:").font = Font(size=10)
        ws.cell(row=row, column=3, value="_________________").font = Font(size=10)

        # ── ШИРИНА КОЛОНОК ──
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 30
        ws.column_dimensions["D"].width = 28
        ws.column_dimensions["E"].width = 10
        ws.column_dimensions["F"].width = 12
        ws.column_dimensions["G"].width = 12
        ws.column_dimensions["H"].width = 14

        # ── ФІКСАЦІЯ ЗАГОЛОВКІВ ──
        ws.freeze_panes = "A5"

        wb.save(filepath)
        return filepath


def calculate_material_order(products: list[dict], project_name: str = "Проєкт") -> MaterialOrder:
    """Швидка функція: розрахувати заявку на матеріали."""
    calc = MaterialCalculator()
    order = calc.calculate(products)
    order.project_name = project_name
    return order


def export_material_order_to_excel(order: MaterialOrder, filepath: str) -> str:
    """Швидка функція: експортувати заявку в Excel."""
    exporter = MaterialOrderExporter(order)
    return exporter.export(filepath)
