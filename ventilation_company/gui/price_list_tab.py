"""🏷️ Вкладка "Прайс-лист" — повноцінний модуль ціноутворення.

Можливості:
  • CRUD для позицій прайсу (власні вироби + перепродаж + послуги)
  • Два прайси: внутрішній (повний) та замовника (публічний)
  • Експорт у PDF, Excel, CSV, HTML
  • Автосинхронізація з вкладкою "Вироби" та "Архів проєктів"
  • Категорії: власне виробництво, перепродаж, монтаж, послуга
"""

from __future__ import annotations

import csv
import io
import json
import os
import tkinter as tk
import zipfile
from collections.abc import Callable
from dataclasses import asdict, dataclass, field
from datetime import datetime
from tkinter import filedialog, messagebox, ttk

# Спробуємо імпортувати openpyxl для Excel
HAVE_OPENPYXL = False
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAVE_OPENPYXL = True
except ImportError:
    pass

# Спробуємо імпортувати reportlab для PDF
HAVE_REPORTLAB = False
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    HAVE_REPORTLAB = True
except ImportError:
    pass


PRICE_LIST_FILE = "data/price_list.json"
ARCHIVE_DIR = "data/archive"


@dataclass
class PriceItem:
    """Одна позиція прайсу."""

    id: str = ""
    name: str = ""
    category: str = "власне виробництво"
    product_type: str = ""
    dimensions: str = ""
    material: str = ""
    thickness: float = 0.0
    unit: str = "шт"
    quantity: int = 1

    cost_price: float = 0.0
    markup_percent: float = 30.0
    labor_cost: float = 0.0
    material_cost: float = 0.0
    overhead_cost: float = 0.0
    supplier: str = ""
    supplier_price: float = 0.0
    notes_internal: str = ""

    unit_price: float = 0.0
    total_price: float = 0.0
    notes_public: str = ""

    created_at: str = field(default_factory=lambda: datetime.now().strftime("%Y-%m-%d %H:%M"))
    updated_at: str = field(default_factory=lambda: datetime.now().strftime("%Y-%m-%d %H:%M"))
    source: str = "manual"  # manual / products / archive
    project_id: str = ""      # ID проєкту, якщо з архіву

    def __post_init__(self):
        if not self.id:
            self.id = f"PRICE-{datetime.now().strftime('%Y%m%d%H%M%S')}-{os.urandom(2).hex()}"
        self.recalculate()

    def recalculate(self):
        if self.category == "перепродаж" and self.supplier_price > 0:
            base = self.supplier_price
        else:
            base = self.cost_price + self.labor_cost + self.overhead_cost
        self.unit_price = base * (1 + self.markup_percent / 100)
        self.total_price = self.unit_price * self.quantity

    @property
    def profit(self) -> float:
        if self.category == "перепродаж" and self.supplier_price > 0:
            base = self.supplier_price
        else:
            base = self.cost_price + self.labor_cost + self.overhead_cost
        return (self.unit_price - base) * self.quantity

    @property
    def display_name(self) -> str:
        """Назва для замовника — без розмірів."""
        if self.product_type:
            return self.product_type.capitalize()
        return self.name

    def to_dict(self) -> dict:
        return asdict(self)

    @classmethod
    def from_dict(cls, data: dict) -> "PriceItem":
        valid_keys = {f.name for f in cls.__dataclass_fields__.values()}
        filtered = {k: v for k, v in data.items() if k in valid_keys}
        return cls(**filtered)


class PriceListManager:
    def __init__(self, filepath: str = PRICE_LIST_FILE):
        self.filepath = filepath
        self.items: list[PriceItem] = []
        self.load()

    def load(self):
        if os.path.exists(self.filepath):
            try:
                with open(self.filepath, encoding="utf-8") as f:
                    data = json.load(f)
                self.items = [PriceItem.from_dict(item) for item in data.get("items", [])]
            except Exception as e:
                print(f"[PriceList] Помилка завантаження: {e}")
                self.items = []
        else:
            self.items = []

    def save(self):
        os.makedirs(os.path.dirname(self.filepath), exist_ok=True)
        with open(self.filepath, "w", encoding="utf-8") as f:
            json.dump(
                {"items": [item.to_dict() for item in self.items]},
                f, ensure_ascii=False, indent=2,
            )

    def add(self, item: PriceItem) -> PriceItem:
        self.items.append(item)
        self.save()
        return item

    def update(self, item_id: str, **kwargs) -> PriceItem | None:
        for item in self.items:
            if item.id == item_id:
                for key, value in kwargs.items():
                    if hasattr(item, key):
                        setattr(item, key, value)
                item.updated_at = datetime.now().strftime("%Y-%m-%d %H:%M")
                item.recalculate()
                self.save()
                return item
        return None

    def delete(self, item_id: str) -> bool:
        for i, item in enumerate(self.items):
            if item.id == item_id:
                del self.items[i]
                self.save()
                return True
        return False

    def get_by_id(self, item_id: str) -> PriceItem | None:
        for item in self.items:
            if item.id == item_id:
                return item
        return None

    def get_by_category(self, category: str) -> list[PriceItem]:
        return [item for item in self.items if item.category == category]

    def get_internal_view(self) -> list[PriceItem]:
        return list(self.items)

    def get_customer_view(self) -> list[PriceItem]:
        return [item for item in self.items if item.unit_price > 0]

    def get_total_internal(self) -> dict:
        total_cost = sum(i.cost_price * i.quantity for i in self.items)
        total_price = sum(i.total_price for i in self.items)
        total_profit = sum(i.profit for i in self.items)
        return {
            "count": len(self.items),
            "total_cost": total_cost,
            "total_price": total_price,
            "total_profit": total_profit,
            "avg_markup": sum(i.markup_percent for i in self.items) / max(len(self.items), 1),
        }

    def clear(self):
        self.items.clear()
        self.save()

    def import_from_products(self, products: list[dict], project_id: str = ""):
        """Імпортувати вироби з вкладки 'Вироби' для конкретного проєкту."""
        imported = 0
        updated = 0

        try:
            from ventilation_company.gui.settings_tab import PricingSettings
            pricing = PricingSettings()
        except Exception:
            pricing = None

        for p in products:
            name = p.get("name", "Виріб")

            # === КОНВЕРТАЦІЯ ТИПІВ (Decimal/Enum → float/str) ===
            def _to_float(val, default=0.0):
                if val is None:
                    return default
                if isinstance(val, (int, float)):
                    return float(val)
                if hasattr(val, "__float__"):  # Decimal
                    return float(val)
                try:
                    return float(str(val).replace(",", "."))
                except (ValueError, TypeError):
                    return default

            def _to_str(val):
                if val is None:
                    return ""
                if isinstance(val, str):
                    return val
                # Enum → .value
                if hasattr(val, "value"):
                    return str(val.value)
                return str(val)

            unit_price = _to_float(p.get("unit_price"), 0)
            quantity = int(_to_float(p.get("quantity"), 1))
            thickness = _to_float(p.get("thickness"), 0)
            metal_area = _to_float(p.get("metal_area_m2"), 0)
            weight = _to_float(p.get("weight_kg"), 0)
            width = _to_float(p.get("width"), 0)
            height = _to_float(p.get("height"), 0)
            length = _to_float(p.get("length"), 0)
            material = _to_str(p.get("material", "оцинкована сталь"))
            product_type = _to_str(p.get("product_type", p.get("type", "")))

            labor = 0.0
            overhead_total = 0.0
            cost_price = 0.0

            if pricing and p.get("metal_area_m2", 0) > 0:
                try:
                    data = {
                        "type": product_type,
                        "material": material,
                        "thickness": thickness,
                        "metal_area_m2": metal_area,
                        "weight_kg": weight,
                        "quantity": quantity,
                        "width": width,
                        "height": height,
                        "length": length,
                        "profile": _to_float(p.get("profile"), 30.0),
                        "angle": _to_float(p.get("angle"), 90),
                        "radius": _to_float(p.get("radius"), 50),
                        "top_extension": _to_float(p.get("top_extension"), 100),
                        "bottom_extension": _to_float(p.get("bottom_extension"), 100),
                    }
                    result = pricing.calculate_product_price_detailed(data)
                    steps = result["steps"]  # list of dicts: {"name": ..., "calc": ..., "value": ...}

                    if len(steps) >= 7:
                        after_waste = steps[1]["value"]
                        after_labor = steps[3]["value"]
                        after_depr = steps[4]["value"]
                        after_elec = steps[5]["value"]
                        after_overhead = steps[6]["value"]
                        final_price = steps[7]["value"] if len(steps) > 7 else after_overhead

                        labor = after_labor - after_waste          # чиста робота
                        depreciation = after_depr - after_labor      # амортизація
                        electricity = after_elec - after_depr        # електроенергія
                        overhead = after_overhead - after_elec       # накладні

                        unit_price = final_price
                        cost_price = after_overhead
                        overhead_total = depreciation + electricity + overhead
                except Exception as e:
                    print(f"[PriceList] PricingSettings failed: {e}, using fallback")

            # Fallback, якщо PricingSettings не спрацював
            if unit_price == 0 and metal_area > 0:
                material_prices = {
                    "оцинкована сталь": {0.5: 260, 0.7: 380, 0.9: 520, 1.0: 750, 1.2: 900},
                    "нержавіюча сталь": {0.5: 350, 0.7: 500, 0.9: 700, 1.0: 1000, 1.2: 1200},
                    "алюміній": {0.5: 200, 0.7: 300, 0.9: 400, 1.0: 550, 1.2: 700},
                }
                price_per_m2 = material_prices.get(material, {}).get(thickness, 260)
                type_coef = {
                    "rect_duct": 1.15, "round_duct": 1.20,
                    "rect_flange": 1.30, "round_flange": 1.30,
                    "rect_tee": 1.50, "round_tee": 1.55,
                    "rect_transition": 1.40, "round_transition": 1.45,
                    "rect_elbow": 1.60, "round_elbow": 1.65,
                    "rect_cap": 1.25, "round_cap": 1.25,
                    "flexible": 1.0,
                }
                coef = type_coef.get(product_type, 1.3)
                unit_price = metal_area * price_per_m2 * coef
                # Розбиваємо fallback на складові
                labor = unit_price * 0.15
                overhead_total = unit_price * 0.10
                cost_price = unit_price - labor - overhead_total

            total_price = unit_price * quantity

            # Формуємо розміри з width/height/length
            dims = f"{width}×{height}×{length}" if length else f"{width}×{height}"

            # Шукаємо існуючий виріб
            existing = None
            for i in self.items:
                if i.name == name and i.source == "products" and i.project_id == str(project_id):
                    existing = i
                    break

            if existing:
                existing.unit_price = unit_price
                existing.total_price = total_price
                existing.cost_price = cost_price
                existing.labor_cost = labor
                existing.overhead_cost = overhead_total
                existing.quantity = quantity
                existing.dimensions = dims
                existing.material = material
                existing.thickness = thickness
                updated += 1
            else:
                item = PriceItem(
                    name=name,
                    category="власне виробництво",
                    product_type=product_type,
                    dimensions=dims,
                    material=material,
                    thickness=thickness,
                    unit="шт",
                    quantity=quantity,
                    cost_price=cost_price,
                    labor_cost=labor,
                    overhead_cost=overhead_total,
                    unit_price=unit_price,
                    total_price=total_price,
                    source="products",
                    project_id=str(project_id),
                )
                self.items.append(item)
                imported += 1

        if imported > 0 or updated > 0:
            self.save()
        return imported

    def import_from_archive(self, project_id: str = None, db_path: str = "data/company.db") -> int:
        """Імпортувати вироби з конкретного проєкту в архіві (SQLite БД)."""
        imported = 0
        if not project_id:
            return 0
        if os.path.exists(db_path):
            try:
                import sqlite3
                conn = sqlite3.connect(db_path)
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()
                cursor.execute("SELECT id, name FROM projects WHERE id = ?", (project_id,))
                project = cursor.fetchone()
                if not project:
                    conn.close()
                    return 0
                project_id_db = project["id"]
                project_name = project["name"]
                cursor.execute("SELECT * FROM project_products WHERE project_id = ?", (project_id_db,))
                products = cursor.fetchall()
                for p in products:
                    item_name = p["name"]
                    if not item_name:
                        continue
                    exists = any(
                        i.name == item_name and i.project_id == str(project_id_db) and i.source == "archive"
                        for i in self.items
                    )
                    if exists:
                        continue
                    w = p["width"] or 0
                    h = p["height"] or 0
                    l = p["length"] or 0
                    dims = f"{w}×{h}×{l}" if l else f"{w}×{h}"
                    unit_price = p["unit_price"] or 0
                    qty = p["quantity"] or 1
                    if unit_price == 0 and p["metal_area_m2"]:
                        material_prices = {"оцинкована сталь": 120.0, "нержавіюча сталь": 350.0, "алюміній": 200.0}
                        area = p["metal_area_m2"] or 0
                        mat = p["material"] or "оцинкована сталь"
                        price_per_m2 = material_prices.get(mat, 120.0)
                        unit_price = area * (price_per_m2 + 50)
                    total_price = unit_price * qty
                    cost = unit_price / 1.3 if unit_price > 0 else 0
                    item = PriceItem(
                        name=item_name,
                        category="власне виробництво",
                        product_type=p["product_type"] or "",
                        dimensions=dims,
                        material=p["material"] or "",
                        thickness=p["thickness"] or 0,
                        unit="шт",
                        quantity=qty,
                        cost_price=cost,
                        labor_cost=0,
                        overhead_cost=0,
                        unit_price=unit_price,
                        total_price=total_price,
                        source="archive",
                        project_id=str(project_id_db),
                        notes_internal=f"З проєкту: {project_name}",
                    )
                    self.items.append(item)
                    imported += 1
                conn.close()
                if imported > 0:
                    self.save()
                return imported
            except Exception as e:
                print(f"[PriceList] Помилка читання з БД: {e}")
        return 0

    def _import_from_zip_archives(self, archive_dir: str = ARCHIVE_DIR) -> int:
        """Імпортувати вироби з ZIP-архівів (старий формат)."""
        imported = 0
        if not os.path.exists(archive_dir):
            return 0

        for filename in os.listdir(archive_dir):
            if not filename.endswith(".zip"):
                continue
            zip_path = os.path.join(archive_dir, filename)
            try:
                with zipfile.ZipFile(zip_path, "r") as zf:
                    for name in zf.namelist():
                        if not name.endswith(".json"):
                            continue
                        data = json.loads(zf.read(name).decode("utf-8"))
                        for p in data.get("products", []):
                            item_name = p.get("name", "")
                            if not item_name:
                                continue
                            exists = any(
                                i.name == item_name and i.source == "archive"
                                for i in self.items
                            )
                            if exists:
                                continue
                            item = PriceItem(
                                name=item_name,
                                category="власне виробництво",
                                product_type=p.get("product_type", ""),
                                dimensions=p.get("dimensions", ""),
                                material=p.get("material", ""),
                                thickness=p.get("thickness", 0),
                                unit="шт",
                                quantity=p.get("quantity", 1),
                                cost_price=p.get("cost_price", 0),
                                unit_price=p.get("unit_price", 0),
                                total_price=p.get("total_price", 0),
                                source="archive",
                                notes_internal=f"З архіву: {filename}",
                            )
                            self.items.append(item)
                            imported += 1
            except Exception as e:
                print(f"[PriceList] Помилка читання {filename}: {e}")
        if imported > 0:
            self.save()
        return imported


class PriceListExporter:
    """Експорт прайс-листа у різні формати."""

    @staticmethod
    def to_csv(items: list[PriceItem], internal: bool = True) -> str:
        output = io.StringIO()
        writer = csv.writer(output)
        if internal:
            writer.writerow([
                "№", "Назва", "Категорія", "Тип", "Розміри", "Матеріал", "Товщ.",
                "Од.", "К-ть", "Собівартість", "Роботи", "Накладні", "Націнка%",
                "Ціна за од.", "Загальна", "Прибуток", "Постачальник", "Примітки"
            ])
            for i, item in enumerate(items, 1):
                writer.writerow([
                    i, item.name, item.category, item.product_type, item.dimensions,
                    item.material, item.thickness, item.unit, item.quantity,
                    f"{item.cost_price:.2f}", f"{item.labor_cost:.2f}",
                    f"{item.overhead_cost:.2f}", f"{item.markup_percent:.1f}",
                    f"{item.unit_price:.2f}", f"{item.total_price:.2f}",
                    f"{item.profit:.2f}", item.supplier, item.notes_internal,
                ])
        else:
            writer.writerow([
                "№", "Назва", "Розміри", "Матеріал", "Товщ.",
                "Од.", "К-ть", "Ціна за од.", "Загальна", "Примітки"
            ])
            for i, item in enumerate(items, 1):
                writer.writerow([
                    i, item.display_name, item.dimensions,
                    item.material, item.thickness, item.unit, item.quantity,
                    f"{item.unit_price:.2f}", f"{item.total_price:.2f}",
                    item.notes_public,
                ])
        return output.getvalue()

    @staticmethod
    def to_excel(items: list[PriceItem], filepath: str, internal: bool = True):
        if not HAVE_OPENPYXL:
            raise ImportError("Встановіть openpyxl: pip install openpyxl")
        wb = Workbook()
        ws = wb.active
        ws.title = "Прайс-лист"
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
        if internal:
            headers = [
                "№", "Назва", "Категорія", "Тип", "Розміри", "Матеріал", "Товщ. (мм)",
                "Од.", "К-ть", "Собівартість", "Роботи", "Накладні", "Націнка %",
                "Ціна за од.", "Загальна", "Прибуток", "Постачальник", "Примітки"
            ]
        else:
            headers = [
                "№", "Назва", "Розміри", "Матеріал", "Товщ. (мм)",
                "Од.", "К-ть", "Ціна за од.", "Загальна", "Примітки"
            ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
        for row, item in enumerate(items, 2):
            if internal:
                values = [
                    row - 1, item.name, item.category, item.product_type,
                    item.dimensions, item.material, item.thickness,
                    item.unit, item.quantity, item.cost_price, item.labor_cost,
                    item.overhead_cost, item.markup_percent, item.unit_price,
                    item.total_price, item.profit, item.supplier, item.notes_internal,
                ]
            else:
                values = [
                    row - 1, item.display_name, item.dimensions,
                    item.material, item.thickness, item.unit, item.quantity,
                    item.unit_price, item.total_price, item.notes_public,
                ]
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except (ValueError, TypeError):
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        total_row = len(items) + 3
        ws.cell(row=total_row, column=1, value="ВСЬОГО:").font = Font(bold=True)
        if internal:
            ws.cell(row=total_row, column=15, value=sum(i.total_price for i in items)).font = Font(bold=True)
            ws.cell(row=total_row, column=16, value=sum(i.profit for i in items)).font = Font(bold=True)
        else:
            ws.cell(row=total_row, column=9, value=sum(i.total_price for i in items)).font = Font(bold=True)
        wb.save(filepath)

    @staticmethod
    def to_pdf(items: list[PriceItem], filepath: str, internal: bool = True, title: str = "Прайс-лист"):
        if not HAVE_REPORTLAB:
            raise ImportError("Встановіть reportlab: pip install reportlab")
        doc = SimpleDocTemplate(
            filepath, pagesize=A4, rightMargin=10 * mm, leftMargin=10 * mm,
            topMargin=15 * mm, bottomMargin=15 * mm
        )
        styles = getSampleStyleSheet()
        story = []
        title_style = ParagraphStyle(
            "CustomTitle", parent=styles["Heading1"], fontSize=16,
            alignment=1, spaceAfter=12, textColor=colors.HexColor("#1565C0")
        )
        story.append(Paragraph(f"<b>{title}</b>", title_style))
        story.append(Paragraph(
            f"Дата формування: {datetime.now().strftime('%d.%m.%Y %H:%M')}",
            styles["Normal"]
        ))
        story.append(Spacer(1, 10))
        if internal:
            headers = ["№", "Назва", "Кат.", "Тип", "Розміри", "Мат.", "Товщ.", "К-ть",
                       "Собіварт.", "Роботи", "Накладні", "Націнка%", "Ціна", "Сума", "Прибуток", "Постач."]
            data = [headers]
            for i, item in enumerate(items, 1):
                data.append([
                    str(i), item.name[:22], item.category[:8], item.product_type[:10],
                    item.dimensions[:12], item.material[:8], str(item.thickness),
                    str(item.quantity), f"{item.cost_price:.2f}", f"{item.labor_cost:.2f}",
                    f"{item.overhead_cost:.2f}", f"{item.markup_percent:.1f}%",
                    f"{item.unit_price:.2f}", f"{item.total_price:.2f}",
                    f"{item.profit:.2f}", item.supplier[:8] or "—",
                ])
            col_widths = [18, 75, 35, 45, 45, 38, 28, 25, 40, 35, 35, 32, 38, 38, 38, 38]
        else:
            headers = ["№", "Назва", "Розміри", "Матеріал", "Товщ.", "К-ть", "Ціна", "Сума"]
            data = [headers]
            for i, item in enumerate(items, 1):
                data.append([
                    str(i), item.display_name[:30], item.dimensions[:18],
                    item.material[:12], str(item.thickness), str(item.quantity),
                    f"{item.unit_price:.2f}", f"{item.total_price:.2f}",
                ])
            col_widths = [22, 110, 70, 60, 35, 35, 55, 55]
        table = Table(data, colWidths=col_widths, repeatRows=1)
        table_style = TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1565C0")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("ALIGN", (1, 1), (1, -1), "LEFT"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("FONTSIZE", (0, 1), (-1, -1), 7),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f5f5")]),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ])
        table.setStyle(table_style)
        story.append(table)
        story.append(Spacer(1, 10))
        total = sum(i.total_price for i in items)
        story.append(Paragraph(f"<b>Всього: {total:,.2f} грн</b>", styles["Heading3"]))
        doc.build(story)

    @staticmethod
    def to_html(items: list[PriceItem], internal: bool = True, title: str = "Прайс-лист") -> str:
        rows = ""
        if internal:
            for i, item in enumerate(items, 1):
                rows += f"""
                <tr>
                    <td>{i}</td><td>{item.name}</td><td>{item.category}</td><td>{item.product_type}</td>
                    <td>{item.dimensions}</td><td>{item.material}</td><td>{item.thickness}</td>
                    <td>{item.unit}</td><td>{item.quantity}</td>
                    <td>{item.cost_price:.2f}</td><td>{item.labor_cost:.2f}</td>
                    <td>{item.overhead_cost:.2f}</td><td>{item.markup_percent:.1f}%</td>
                    <td>{item.unit_price:.2f}</td><td>{item.total_price:.2f}</td>
                    <td>{item.profit:.2f}</td><td>{item.supplier}</td><td>{item.notes_internal}</td>
                </tr>"""
            headers = """
                <th>№</th><th>Назва</th><th>Категорія</th><th>Тип</th><th>Розміри</th>
                <th>Матеріал</th><th>Товщ.</th><th>Од.</th><th>К-ть</th>
                <th>Собіварт.</th><th>Роботи</th><th>Накладні</th><th>Націнка</th>
                <th>Ціна од.</th><th>Сума</th><th>Прибуток</th><th>Постач.</th><th>Примітки</th>
            """
        else:
            for i, item in enumerate(items, 1):
                rows += f"""
                <tr>
                    <td>{i}</td><td>{item.display_name}</td>
                    <td>{item.dimensions}</td><td>{item.material}</td><td>{item.thickness}</td>
                    <td>{item.unit}</td><td>{item.quantity}</td>
                    <td>{item.unit_price:.2f}</td><td>{item.total_price:.2f}</td>
                    <td>{item.notes_public}</td>
                </tr>"""
            headers = """
                <th>№</th><th>Назва</th><th>Розміри</th>
                <th>Матеріал</th><th>Товщ.</th><th>Од.</th><th>К-ть</th>
                <th>Ціна од.</th><th>Сума</th><th>Примітки</th>
            """
        total = sum(i.total_price for i in items)
        total_profit = sum(i.profit for i in items) if internal else 0
        return f"""<!DOCTYPE html>
<html lang="uk">
<head>
    <meta charset="UTF-8">
    <title>{title}</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 20px; }}
        h1 {{ color: #1565C0; text-align: center; }}
        .meta {{ text-align: center; color: #666; margin-bottom: 20px; }}
        table {{ border-collapse: collapse; width: 100%; font-size: 11px; }}
        th {{ background: #1565C0; color: white; padding: 8px; text-align: center; }}
        td {{ border: 1px solid #ccc; padding: 6px; text-align: center; }}
        tr:nth-child(even) {{ background: #f5f5f5; }}
        .total {{ font-weight: bold; font-size: 14px; margin-top: 15px; text-align: right; }}
        .profit {{ color: #2E7D32; }}
        @media print {{ body {{ margin: 0; }} }}
    </style>
</head>
<body>
    <h1>🏷️ {title}</h1>
    <p class="meta">Дата: {datetime.now().strftime('%d.%m.%Y %H:%M')}</p>
    <table>
        <thead><tr>{headers}</tr></thead>
        <tbody>{rows}</tbody>
    </table>
    <p class="total">Всього: <b>{total:,.2f} грн</b></p>
    {f'<p class="total profit">Прибуток: <b>{total_profit:,.2f} грн</b></p>' if internal else ''}
</body>
</html>"""


class PriceListTab:
    """Вкладка прайс-листа."""

    CATEGORIES = ["власне виробництво", "перепродаж", "монтаж", "послуга"]
    UNITS = ["шт", "м", "м²", "м³", "кг", "комплект"]

    def __init__(self, parent: ttk.Notebook, get_products_callback: Callable | None = None):
        self.frame = ttk.Frame(parent)
        self.manager = PriceListManager()
        self.get_products_callback = get_products_callback
        self._current_view = "internal"
        self._selected_item_id: str | None = None
        self._build_ui()
        self._refresh_tree()

    def _build_ui(self):
        top = ttk.Frame(self.frame, padding=5)
        top.pack(fill=tk.X)

        ttk.Label(top, text="🏷️ Прайс-лист", font=("Arial", 14, "bold")).pack(side=tk.LEFT)

        btn_frame = ttk.Frame(top)
        btn_frame.pack(side=tk.LEFT, padx=(20, 0))

        ttk.Button(btn_frame, text="➕ Додати", command=self._add_dialog).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="✏️ Редагувати", command=self._edit_dialog).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="🗑️ Видалити", command=self._delete_selected).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="📋 Дублювати", command=self._duplicate_selected).pack(side=tk.LEFT, padx=2)

        # Синхронізація
        sync_frame = ttk.LabelFrame(top, text="Синхронізація", padding=3)
        sync_frame.pack(side=tk.LEFT, padx=(10, 0))
        ttk.Button(sync_frame, text="🔄 З виробів", command=self._sync_from_products).pack(side=tk.LEFT, padx=2)
        ttk.Button(sync_frame, text="📦 З архіву", command=self._sync_from_archive).pack(side=tk.LEFT, padx=2)
        ttk.Button(sync_frame, text="♻️ Оновити прайс", command=self._refresh_current_project).pack(side=tk.LEFT, padx=2)

        # Експорт
        export_frame = ttk.LabelFrame(top, text="Експорт", padding=3)
        export_frame.pack(side=tk.RIGHT, padx=5)
        ttk.Button(export_frame, text="📄 PDF", command=lambda: self._export("pdf")).pack(side=tk.LEFT, padx=2)
        ttk.Button(export_frame, text="📊 Excel", command=lambda: self._export("excel")).pack(side=tk.LEFT, padx=2)
        ttk.Button(export_frame, text="🌐 HTML", command=lambda: self._export("html")).pack(side=tk.LEFT, padx=2)
        ttk.Button(export_frame, text="📋 CSV", command=lambda: self._export("csv")).pack(side=tk.LEFT, padx=2)
        ttk.Button(export_frame, text="🖨️ Друк", command=self._print_dialog).pack(side=tk.LEFT, padx=2)

        view_frame = ttk.LabelFrame(self.frame, text="Режим перегляду", padding=5)
        view_frame.pack(fill=tk.X, padx=5, pady=(5, 0))

        self.view_var = tk.StringVar(value="internal")
        ttk.Radiobutton(
            view_frame, text="🔐 Внутрішній прайс (повна інформація)",
            variable=self.view_var, value="internal", command=self._on_view_changed
        ).pack(side=tk.LEFT, padx=10)
        ttk.Radiobutton(
            view_frame, text="📋 Прайс замовника (публічний)",
            variable=self.view_var, value="customer", command=self._on_view_changed
        ).pack(side=tk.LEFT, padx=10)

        filter_frame = ttk.Frame(self.frame, padding=5)
        filter_frame.pack(fill=tk.X, padx=5)

        ttk.Label(filter_frame, text="Фільтр категорії:").pack(side=tk.LEFT)
        self.filter_cat_var = tk.StringVar(value="всі")
        ttk.Combobox(
            filter_frame, textvariable=self.filter_cat_var,
            values=["всі"] + self.CATEGORIES, state="readonly", width=20
        ).pack(side=tk.LEFT, padx=5)
        self.filter_cat_var.trace_add("write", lambda *args: self._refresh_tree())

        ttk.Label(filter_frame, text="Пошук:").pack(side=tk.LEFT, padx=(15, 0))
        self.search_var = tk.StringVar()
        ttk.Entry(filter_frame, textvariable=self.search_var, width=25).pack(side=tk.LEFT, padx=5)
        ttk.Button(filter_frame, text="🔍", width=3, command=self._refresh_tree).pack(side=tk.LEFT)
        self.search_var.trace_add("write", lambda *args: self._refresh_tree())

        table_frame = ttk.Frame(self.frame)
        table_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        self.internal_columns = (
            "num", "name", "category", "type", "dimensions", "material", "thickness",
            "unit", "qty", "cost", "labor", "overhead", "markup",
            "unit_price", "total", "profit", "supplier", "notes"
        )
        self.internal_headings = {
            "num": "№", "name": "Назва", "category": "Категорія", "type": "Тип",
            "dimensions": "Розміри", "material": "Матеріал", "thickness": "Товщ.",
            "unit": "Од.", "qty": "К-ть", "cost": "Собіварт.", "labor": "Роботи",
            "overhead": "Накладні", "markup": "Націнка%", "unit_price": "Ціна од.",
            "total": "Сума", "profit": "Прибуток", "supplier": "Постач.", "notes": "Примітки"
        }
        self.internal_widths = {
            "num": 30, "name": 150, "category": 90, "type": 100, "dimensions": 90,
            "material": 90, "thickness": 45, "unit": 40, "qty": 45, "cost": 70,
            "labor": 60, "overhead": 60, "markup": 55, "unit_price": 70,
            "total": 80, "profit": 70, "supplier": 90, "notes": 100
        }

        # === ЗАМОВНИК: без колонки "Тип" ===
        self.customer_columns = (
            "num", "name", "dimensions", "material", "thickness",
            "unit", "qty", "unit_price", "total", "notes"
        )
        self.customer_headings = {
            "num": "№", "name": "Назва", "dimensions": "Розміри",
            "material": "Матеріал", "thickness": "Товщ.", "unit": "Од.",
            "qty": "К-ть", "unit_price": "Ціна за од.", "total": "Загальна", "notes": "Примітки"
        }
        self.customer_widths = {
            "num": 35, "name": 200, "dimensions": 120,
            "material": 100, "thickness": 50, "unit": 45, "qty": 50,
            "unit_price": 90, "total": 90, "notes": 150
        }

        self.tree = ttk.Treeview(table_frame, show="headings", height=20)
        self._setup_tree_columns()

        vsb = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=self.tree.yview)
        hsb = ttk.Scrollbar(table_frame, orient=tk.HORIZONTAL, command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)

        self.tree.bind("<Double-1>", lambda e: self._edit_dialog())
        self.tree.bind("<Button-3>", self._on_right_click)

        self.summary_frame = ttk.LabelFrame(self.frame, text="Зведення", padding=5)
        self.summary_frame.pack(fill=tk.X, padx=5, pady=(0, 5))

        self.summary_label = ttk.Label(self.summary_frame, text="", font=("Consolas", 10))
        self.summary_label.pack(anchor=tk.W)

    def _setup_tree_columns(self):
        for col in self.tree["columns"]:
            self.tree.heading(col, text="")
        self.tree.delete(*self.tree.get_children())

        if self._current_view == "internal":
            cols = self.internal_columns
            headings = self.internal_headings
            widths = self.internal_widths
        else:
            cols = self.customer_columns
            headings = self.customer_headings
            widths = self.customer_widths

        self.tree["columns"] = cols
        for col in cols:
            self.tree.heading(col, text=headings.get(col, col))
            self.tree.column(col, width=widths.get(col, 80), anchor=tk.CENTER if col != "name" else tk.W)

    def _on_view_changed(self):
        self._current_view = self.view_var.get()
        self._setup_tree_columns()
        self._refresh_tree()

    def _get_filtered_items(self) -> list[PriceItem]:
        if self._current_view == "internal":
            items = self.manager.get_internal_view()
        else:
            items = self.manager.get_customer_view()

        cat = self.filter_cat_var.get()
        if cat != "всі":
            items = [i for i in items if i.category == cat]

        search = self.search_var.get().lower().strip()
        if search:
            items = [
                i for i in items
                if search in i.name.lower()
                or search in i.product_type.lower()
                or search in i.dimensions.lower()
                or search in i.material.lower()
                or search in i.supplier.lower()
            ]
        return items

    def _refresh_tree(self):
        for item in self.tree.get_children():
            self.tree.delete(item)

        items = self._get_filtered_items()

        for i, item in enumerate(items, 1):
            if self._current_view == "internal":
                values = (
                    i, item.name, item.category, item.product_type, item.dimensions,
                    item.material, item.thickness, item.unit, item.quantity,
                    f"{item.cost_price:.2f}", f"{item.labor_cost:.2f}",
                    f"{item.overhead_cost:.2f}", f"{item.markup_percent:.1f}",
                    f"{item.unit_price:.2f}", f"{item.total_price:.2f}",
                    f"{item.profit:.2f}", item.supplier, item.notes_internal,
                )
            else:
                # === ЗАМОВНИК: назва без розмірів (product_type), без колонки "Тип" ===
                values = (
                    i, item.display_name, item.dimensions,
                    item.material, item.thickness, item.unit, item.quantity,
                    f"{item.unit_price:.2f}", f"{item.total_price:.2f}",
                    item.notes_public,
                )
            self.tree.insert("", tk.END, values=values, tags=(item.id,))

        self._update_summary(items)

    def _update_summary(self, items: list[PriceItem]):
        total_qty = sum(i.quantity for i in items)
        total_price = sum(i.total_price for i in items)

        if self._current_view == "internal":
            total_cost = sum(i.cost_price * i.quantity for i in items)
            total_labor = sum(i.labor_cost * i.quantity for i in items)
            total_overhead = sum(i.overhead_cost * i.quantity for i in items)
            total_profit = sum(i.profit for i in items)
            text = (
                f"Позицій: {len(items)}  |  К-ть: {total_qty}  |  "
                f"Собівартість: {total_cost:,.2f} грн  |  Роботи: {total_labor:,.2f} грн  |  "
                f"Накладні: {total_overhead:,.2f} грн  |  Загальна: {total_price:,.2f} грн  |  "
                f"Прибуток: {total_profit:,.2f} грн"
            )
        else:
            text = f"Позицій: {len(items)}  |  К-ть: {total_qty}  |  Загальна: {total_price:,.2f} грн"

        self.summary_label.config(text=text)

    def _get_selected_item(self) -> PriceItem | None:
        selected = self.tree.selection()
        if not selected:
            return None
        idx = self.tree.index(selected[0])
        items = self._get_filtered_items()
        if 0 <= idx < len(items):
            return items[idx]
        return None

    def _add_dialog(self):
        self._open_item_dialog()

    def _edit_dialog(self):
        item = self._get_selected_item()
        if not item:
            messagebox.showwarning("Увага", "Оберіть позицію для редагування")
            return
        self._open_item_dialog(item)

    def _open_item_dialog(self, item: PriceItem | None = None):
        dialog = tk.Toplevel(self.frame)
        dialog.title("Редагувати позицію" if item else "Додати позицію в прайс")
        dialog.geometry("500x650")
        dialog.minsize(400, 300)
        dialog.resizable(True, True)
        dialog.transient(self.frame)
        dialog.grab_set()

        is_edit = item is not None

        vars_dict = {
            "name": tk.StringVar(value=item.name if is_edit else ""),
            "category": tk.StringVar(value=item.category if is_edit else "власне виробництво"),
            "product_type": tk.StringVar(value=item.product_type if is_edit else ""),
            "dimensions": tk.StringVar(value=item.dimensions if is_edit else ""),
            "material": tk.StringVar(value=item.material if is_edit else ""),
            "thickness": tk.StringVar(value=str(item.thickness) if is_edit else "0.7"),
            "unit": tk.StringVar(value=item.unit if is_edit else "шт"),
            "quantity": tk.StringVar(value=str(item.quantity) if is_edit else "1"),
            "cost_price": tk.StringVar(value=str(item.cost_price) if is_edit else "0"),
            "labor_cost": tk.StringVar(value=str(item.labor_cost) if is_edit else "0"),
            "overhead_cost": tk.StringVar(value=str(item.overhead_cost) if is_edit else "0"),
            "markup_percent": tk.StringVar(value=str(item.markup_percent) if is_edit else "30"),
            "supplier": tk.StringVar(value=item.supplier if is_edit else ""),
            "supplier_price": tk.StringVar(value=str(item.supplier_price) if is_edit else "0"),
            "notes_internal": tk.StringVar(value=item.notes_internal if is_edit else ""),
            "notes_public": tk.StringVar(value=item.notes_public if is_edit else ""),
        }

        row = 0
        def add_row(label_text, var, entry_width=15):
            nonlocal row
            ttk.Label(dialog, text=label_text).grid(row=row, column=0, sticky=tk.W, padx=10, pady=2)
            ttk.Entry(dialog, textvariable=var, width=entry_width).grid(row=row, column=1, sticky=tk.W, padx=5, pady=2)
            row += 1

        add_row("Назва *:", vars_dict["name"], 35)

        ttk.Label(dialog, text="Категорія:").grid(row=row, column=0, sticky=tk.W, padx=10, pady=2)
        ttk.Combobox(dialog, textvariable=vars_dict["category"], values=self.CATEGORIES, state="readonly", width=20).grid(row=row, column=1, sticky=tk.W, padx=5, pady=2)
        row += 1

        add_row("Тип виробу:", vars_dict["product_type"], 25)
        add_row("Розміри:", vars_dict["dimensions"], 25)
        add_row("Матеріал:", vars_dict["material"], 20)
        add_row("Товщина (мм):", vars_dict["thickness"])

        ttk.Label(dialog, text="Од. виміру:").grid(row=row, column=0, sticky=tk.W, padx=10, pady=2)
        ttk.Combobox(dialog, textvariable=vars_dict["unit"], values=self.UNITS, state="readonly", width=10).grid(row=row, column=1, sticky=tk.W, padx=5, pady=2)
        row += 1

        add_row("Кількість:", vars_dict["quantity"])

        ttk.Separator(dialog, orient=tk.HORIZONTAL).grid(row=row, column=0, columnspan=2, sticky="ew", pady=10)
        row += 1
        ttk.Label(dialog, text="💰 Фінанси (внутрішні)", font=("Arial", 10, "bold")).grid(row=row, column=0, columnspan=2, sticky=tk.W, padx=10, pady=5)
        row += 1

        add_row("Собівартість за од.:", vars_dict["cost_price"])
        add_row("Вартість робіт за од.:", vars_dict["labor_cost"])
        add_row("Накладні витрати за од.:", vars_dict["overhead_cost"])
        add_row("Націнка (%):", vars_dict["markup_percent"])

        ttk.Separator(dialog, orient=tk.HORIZONTAL).grid(row=row, column=0, columnspan=2, sticky="ew", pady=10)
        row += 1
        ttk.Label(dialog, text="🔄 Перепродаж", font=("Arial", 10, "bold")).grid(row=row, column=0, columnspan=2, sticky=tk.W, padx=10, pady=5)
        row += 1

        add_row("Постачальник:", vars_dict["supplier"], 25)
        add_row("Закупівельна ціна:", vars_dict["supplier_price"])

        ttk.Separator(dialog, orient=tk.HORIZONTAL).grid(row=row, column=0, columnspan=2, sticky="ew", pady=10)
        row += 1
        ttk.Label(dialog, text="📝 Примітки", font=("Arial", 10, "bold")).grid(row=row, column=0, columnspan=2, sticky=tk.W, padx=10, pady=5)
        row += 1

        add_row("Внутрішні:", vars_dict["notes_internal"], 35)
        add_row("Публічні:", vars_dict["notes_public"], 35)

        btn_frame = ttk.Frame(dialog)
        btn_frame.grid(row=row, column=0, columnspan=2, pady=15)

        def save():
            try:
                name = vars_dict["name"].get().strip()
                if not name:
                    messagebox.showwarning("Увага", "Вкажіть назву позиції")
                    return

                qty = max(1, int(float(vars_dict["quantity"].get() or 1)))
                thickness = float(vars_dict["thickness"].get() or 0)
                cost = float(vars_dict["cost_price"].get() or 0)
                labor = float(vars_dict["labor_cost"].get() or 0)
                overhead = float(vars_dict["overhead_cost"].get() or 0)
                markup = float(vars_dict["markup_percent"].get() or 30)
                supplier_price = float(vars_dict["supplier_price"].get() or 0)

                if is_edit:
                    self.manager.update(
                        item.id,
                        name=name,
                        category=vars_dict["category"].get(),
                        product_type=vars_dict["product_type"].get(),
                        dimensions=vars_dict["dimensions"].get(),
                        material=vars_dict["material"].get(),
                        thickness=thickness,
                        unit=vars_dict["unit"].get(),
                        quantity=qty,
                        cost_price=cost,
                        labor_cost=labor,
                        overhead_cost=overhead,
                        markup_percent=markup,
                        supplier=vars_dict["supplier"].get(),
                        supplier_price=supplier_price,
                        notes_internal=vars_dict["notes_internal"].get(),
                        notes_public=vars_dict["notes_public"].get(),
                    )
                else:
                    new_item = PriceItem(
                        name=name,
                        category=vars_dict["category"].get(),
                        product_type=vars_dict["product_type"].get(),
                        dimensions=vars_dict["dimensions"].get(),
                        material=vars_dict["material"].get(),
                        thickness=thickness,
                        unit=vars_dict["unit"].get(),
                        quantity=qty,
                        cost_price=cost,
                        labor_cost=labor,
                        overhead_cost=overhead,
                        markup_percent=markup,
                        supplier=vars_dict["supplier"].get(),
                        supplier_price=supplier_price,
                        notes_internal=vars_dict["notes_internal"].get(),
                        notes_public=vars_dict["notes_public"].get(),
                    )
                    self.manager.add(new_item)

                self._refresh_tree()
                dialog.destroy()
            except ValueError as e:
                messagebox.showwarning("Увага", f"Помилка в даних: {e}")

        ttk.Button(btn_frame, text="✅ Зберегти", command=save).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="❌ Скасувати", command=dialog.destroy).pack(side=tk.LEFT, padx=5)

    def _delete_selected(self):
        item = self._get_selected_item()
        if not item:
            messagebox.showwarning("Увага", "Оберіть позицію для видалення")
            return
        if messagebox.askyesno("Підтвердження", f'Видалити "{item.name}"?'):
            self.manager.delete(item.id)
            self._refresh_tree()

    def _duplicate_selected(self):
        item = self._get_selected_item()
        if not item:
            messagebox.showwarning("Увага", "Оберіть позицію для дублювання")
            return
        new_item = PriceItem(
            name=f"{item.name} (копія)",
            category=item.category,
            product_type=item.product_type,
            dimensions=item.dimensions,
            material=item.material,
            thickness=item.thickness,
            unit=item.unit,
            quantity=item.quantity,
            cost_price=item.cost_price,
            labor_cost=item.labor_cost,
            overhead_cost=item.overhead_cost,
            markup_percent=item.markup_percent,
            supplier=item.supplier,
            supplier_price=item.supplier_price,
            notes_internal=item.notes_internal,
            notes_public=item.notes_public,
        )
        self.manager.add(new_item)
        self._refresh_tree()

    def _sync_from_products(self):
        """Синхронізувати вироби з вкладки 'Вироби' для поточного проєкту."""
        project_id = getattr(self, '_current_project_id', '') or 'current'
        if self.get_products_callback:
            products = self.get_products_callback()
            if products:
                count = self.manager.import_from_products(products, project_id=project_id)
                self._refresh_tree()
                if count > 0:
                    messagebox.showinfo("Синхронізація", f"Імпортовано {count} нових позицій з виробів")
                else:
                    messagebox.showinfo("Синхронізація", "Усі вироби вже в прайсі")
            else:
                messagebox.showinfo("Синхронізація", "Немає виробів для імпорту")
        else:
            messagebox.showwarning("Увага", "Функція синхронізації з виробами недоступна")

    def _sync_from_archive(self):
        """Синхронізувати з конкретного проєкту в архіві."""
        project_id = getattr(self, '_current_project_id', None)
        if not project_id:
            try:
                from ventilation_company.db_integration import ProjectDatabase
                db = ProjectDatabase("data/company.db")
                projects = db.list_projects()
                if not projects:
                    messagebox.showinfo("Архів", "Архів проєктів порожній")
                    return
                dialog = tk.Toplevel(self.frame)
                dialog.title("Виберіть проєкт")
                dialog.geometry("400x300")
                dialog.minsize(400, 300)
                dialog.resizable(True, True)
                dialog.transient(self.frame)
                dialog.grab_set()
                listbox = tk.Listbox(dialog, font=("Arial", 11))
                listbox.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
                project_map = {}
                for p in projects:
                    display = f"{p.get('name', 'Без назви')} (ID: {p.get('id', '?')})"
                    project_map[display] = p.get('id')
                    listbox.insert(tk.END, display)
                def on_select():
                    sel = listbox.curselection()
                    if not sel:
                        return
                    selected_id = project_map[listbox.get(sel[0])]
                    dialog.destroy()
                    self._do_archive_sync(selected_id)
                ttk.Button(dialog, text="Імпортувати", command=on_select).pack(pady=5)
            except Exception as e:
                messagebox.showerror("Помилка", f"Не вдалося відкрити архів: {e}")
        else:
            self._do_archive_sync(project_id)

    def _do_archive_sync(self, project_id: str):
        """Виконати імпорт з архіву для конкретного проєкту."""
        count = self.manager.import_from_archive(project_id=project_id)
        self._refresh_tree()
        if count > 0:
            messagebox.showinfo("Синхронізація", f"Імпортовано {count} нових позицій з архіву")
        else:
            messagebox.showinfo("Синхронізація", "Усі позиції вже синхронізовані або проєкт порожній")

    def _refresh_current_project(self):
        """Оновити прайс-лист для поточного проєкту (перезавантажити дані)."""
        project_id = getattr(self, '_current_project_id', None)
        if not project_id:
            messagebox.showwarning("Увага", "Спочатку відкрийте або створіть проєкт")
            return
        old_count = len(self.manager.items)
        self.manager.items = [
            i for i in self.manager.items
            if not (i.source in ("products", "archive") and i.project_id == str(project_id))
        ]
        removed = old_count - len(self.manager.items)
        self.manager.save()
        self._sync_from_products()
        self._do_archive_sync(project_id)
        self._refresh_tree()
        messagebox.showinfo("Оновлення", f"Прайс оновлено. Видалено старих: {removed}")

    def _export(self, fmt: str):
        items = self._get_filtered_items()
        if not items:
            messagebox.showwarning("Увага", "Немає даних для експорту")
            return

        internal = self._current_view == "internal"

        try:
            if fmt == "csv":
                filepath = filedialog.asksaveasfilename(
                    defaultextension=".csv", filetypes=[("CSV", "*.csv")],
                    initialfile=f"price_list_{datetime.now().strftime('%Y%m%d')}.csv"
                )
                if filepath:
                    content = PriceListExporter.to_csv(items, internal)
                    with open(filepath, "w", encoding="utf-8-sig") as f:
                        f.write(content)
                    messagebox.showinfo("Успіх", f"Збережено: {filepath}")

            elif fmt == "excel":
                if not HAVE_OPENPYXL:
                    messagebox.showwarning("Увага", "Встановіть openpyxl: pip install openpyxl")
                    return
                filepath = filedialog.asksaveasfilename(
                    defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")],
                    initialfile=f"price_list_{datetime.now().strftime('%Y%m%d')}.xlsx"
                )
                if filepath:
                    PriceListExporter.to_excel(items, filepath, internal)
                    messagebox.showinfo("Успіх", f"Збережено: {filepath}")

            elif fmt == "pdf":
                if not HAVE_REPORTLAB:
                    messagebox.showwarning("Увага", "Встановіть reportlab: pip install reportlab")
                    return
                filepath = filedialog.asksaveasfilename(
                    defaultextension=".pdf", filetypes=[("PDF", "*.pdf")],
                    initialfile=f"price_list_{datetime.now().strftime('%Y%m%d')}.pdf"
                )
                if filepath:
                    title = "Прайс-лист (внутрішній)" if internal else "Прайс-лист для замовника"
                    PriceListExporter.to_pdf(items, filepath, internal, title)
                    messagebox.showinfo("Успіх", f"Збережено: {filepath}")

            elif fmt == "html":
                filepath = filedialog.asksaveasfilename(
                    defaultextension=".html", filetypes=[("HTML", "*.html")],
                    initialfile=f"price_list_{datetime.now().strftime('%Y%m%d')}.html"
                )
                if filepath:
                    title = "Прайс-лист (внутрішній)" if internal else "Прайс-лист для замовника"
                    content = PriceListExporter.to_html(items, internal, title)
                    with open(filepath, "w", encoding="utf-8") as f:
                        f.write(content)
                    messagebox.showinfo("Успіх", f"Збережено: {filepath}")

        except Exception as e:
            messagebox.showerror("Помилка", str(e))

    def _print_dialog(self):
        items = self._get_filtered_items()
        if not items:
            messagebox.showwarning("Увага", "Немає даних для друку")
            return

        internal = self._current_view == "internal"
        title = "Прайс-лист (внутрішній)" if internal else "Прайс-лист для замовника"
        content = PriceListExporter.to_html(items, internal, title)

        import tempfile
        with tempfile.NamedTemporaryFile("w", suffix=".html", delete=False, encoding="utf-8") as f:
            f.write(content)
            temp_path = f.name

        import webbrowser
        webbrowser.open(f"file:///{temp_path}")

    def _on_right_click(self, event):
        item = self.tree.identify_row(event.y)
        if item:
            self.tree.selection_set(item)
            menu = tk.Menu(self.frame, tearoff=0)
            menu.add_command(label="✏️ Редагувати", command=self._edit_dialog)
            menu.add_command(label="📋 Дублювати", command=self._duplicate_selected)
            menu.add_separator()
            menu.add_command(label="🗑️ Видалити", command=self._delete_selected)
            menu.post(event.x_root, event.y_root)

    def get_manager(self) -> PriceListManager:
        return self.manager

    def get_items(self) -> list[PriceItem]:
        return self.manager.items
