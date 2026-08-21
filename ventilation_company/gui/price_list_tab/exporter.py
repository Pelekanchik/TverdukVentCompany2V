"""🏷️ Вкладка "Прайс-лист" — повноцінний модуль ціноутворення.

Можливості:
  • CRUD для позицій прайсу (власні вироби + перепродаж + послуги)
  • Два прайси: внутрішній (повний) та замовника (публічний)
  • Експорт у PDF, Excel, CSV, HTML
  • Автосинхронізація з вкладкою "Вироби" та "Архів проєктів"
  • Категорії: власне виробництво, перепродаж, монтаж, послуга
"""

from __future__ import annotations

import csv
import pathlib
import io
import json
import os
import tkinter as tk
import zipfile
from collections.abc import Callable
from dataclasses import asdict, dataclass, field
from datetime import datetime
from tkinter import filedialog, messagebox, ttk

# Спробуємо імпортувати openpyxl для Excel
HAVE_OPENPYXL = False
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAVE_OPENPYXL = True
except ImportError:
    pass

# Спробуємо імпортувати reportlab для PDF
HAVE_REPORTLAB = False
_PDF_FONT_NAME = "Helvetica"
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    # Реєструємо системний шрифт з підтримкою кирилиці
    for _fp in [
        r"C:\Windows\Fonts\arial.ttf",
        r"C:\Windows\Fonts\calibri.ttf",
        r"C:\Windows\Fonts\segoeui.ttf",
        r"C:\Windows\Fonts\tahoma.ttf",
    ]:
        if pathlib.Path(_fp).exists():
            try:
                _fname = pathlib.Path(_fp).stem.capitalize()
                pdfmetrics.registerFont(TTFont(_fname, _fp))
                pdfmetrics.registerFont(TTFont(_fname + "-Bold", _fp))
                _PDF_FONT_NAME = _fname
                break
            except Exception:
                continue

    HAVE_REPORTLAB = True
except ImportError:
    pass


from ventilation_company.gui.price_list_tab.models import PriceItem


PRICE_LIST_FILE = "data/price_list.json"
ARCHIVE_DIR = "data/archive"


class PriceListExporter:
    """Експорт прайс-листа у різні формати."""

    @staticmethod
    def to_csv(items: list[PriceItem], internal: bool = True) -> str:
        output = io.StringIO()
        writer = csv.writer(output)
        if internal:
            writer.writerow([
                "№", "Назва", "Категорія", "Тип", "Розміри", "Матеріал", "Товщ.",
                "Од.", "К-ть", "Собівартість", "Роботи", "Накладні", "Націнка%",
                "Ціна за од.", "Загальна", "Прибуток", "Постачальник", "Примітки"
            ])
            for i, item in enumerate(items, 1):
                writer.writerow([
                    i, item.name, item.category, item.product_type, item.dimensions,
                    item.material, item.thickness, item.unit, item.quantity,
                    f"{item.cost_price:.2f}", f"{item.labor_cost:.2f}",
                    f"{item.overhead_cost:.2f}", f"{item.markup_percent:.1f}",
                    f"{item.unit_price:.2f}", f"{item.total_price:.2f}",
                    f"{item.profit:.2f}", item.supplier, item.notes_internal,
                ])
        else:
            writer.writerow([
                "№", "Назва", "Тип", "Розміри", "Матеріал", "Товщ.",
                "Од.", "К-ть", "Ціна за од.", "Загальна", "Примітки"
            ])
            for i, item in enumerate(items, 1):
                writer.writerow([
                    i, item.name, item.product_type, item.dimensions,
                    item.material, item.thickness, item.unit, item.quantity,
                    f"{item.unit_price:.2f}", f"{item.total_price:.2f}",
                    item.notes_public,
                ])
        return output.getvalue()

    @staticmethod
    def to_excel(items: list[PriceItem], filepath: str, internal: bool = True):
        if not HAVE_OPENPYXL:
            raise ImportError("Встановіть openpyxl: pip install openpyxl")
        wb = Workbook()
        ws = wb.active
        ws.title = "Прайс-лист"
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
        if internal:
            headers = [
                "№", "Назва", "Категорія", "Тип", "Розміри", "Матеріал", "Товщ. (мм)",
                "Од.", "К-ть", "Собівартість", "Роботи", "Накладні", "Націнка %",
                "Ціна за од.", "Загальна", "Прибуток", "Постачальник", "Примітки"
            ]
        else:
            headers = [
                "№", "Назва", "Тип", "Розміри", "Матеріал", "Товщ. (мм)",
                "Од.", "К-ть", "Ціна за од.", "Загальна", "Примітки"
            ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
        for row, item in enumerate(items, 2):
            if internal:
                values = [
                    row - 1, item.name, item.category, item.product_type,
                    item.dimensions, item.material, item.thickness,
                    item.unit, item.quantity, item.cost_price, item.labor_cost,
                    item.overhead_cost, item.markup_percent, item.unit_price,
                    item.total_price, item.profit, item.supplier, item.notes_internal,
                ]
            else:
                values = [
                    row - 1, item.name, item.product_type, item.dimensions,
                    item.material, item.thickness, item.unit, item.quantity,
                    item.unit_price, item.total_price, item.notes_public,
                ]
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except (ValueError, TypeError):
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        total_row = len(items) + 3
        ws.cell(row=total_row, column=1, value="ВСЬОГО:").font = Font(bold=True)
        if internal:
            ws.cell(row=total_row, column=15, value=sum(i.total_price for i in items)).font = Font(bold=True)
            ws.cell(row=total_row, column=16, value=sum(i.profit for i in items)).font = Font(bold=True)
        else:
            ws.cell(row=total_row, column=10, value=sum(i.total_price for i in items)).font = Font(bold=True)
        wb.save(filepath)

    @staticmethod
    def to_pdf(items: list[PriceItem], filepath: str, internal: bool = True, title: str = "Прайс-лист"):
        if not HAVE_REPORTLAB:
            raise ImportError("Встановіть reportlab: pip install reportlab")
        doc = SimpleDocTemplate(
            filepath, pagesize=A4, rightMargin=15 * mm, leftMargin=15 * mm,
            topMargin=15 * mm, bottomMargin=15 * mm
        )
        styles = getSampleStyleSheet()
        story = []
        title_style = ParagraphStyle(
            "CustomTitle", parent=styles["Heading1"], fontSize=16,
            alignment=1, spaceAfter=12, textColor=colors.HexColor("#1565C0"),
            fontName=_PDF_FONT_NAME
        )
        normal_style = ParagraphStyle(
            "CustomNormal", parent=styles["Normal"], fontName=_PDF_FONT_NAME
        )
        story.append(Paragraph(f"<b>{title}</b>", title_style))
        story.append(Paragraph(
            f"Дата формування: {datetime.now().strftime('%d.%m.%Y %H:%M')}",
            normal_style
        ))
        story.append(Spacer(1, 10))
        if internal:
            headers = ["№", "Назва", "Кат.", "Тип", "Розміри", "Мат.", "Товщ.", "К-ть", "Ціна", "Сума"]
            data = [headers]
            for i, item in enumerate(items, 1):
                data.append([
                    str(i), item.name[:25], item.category[:8], item.product_type[:12],
                    item.dimensions[:15], item.material[:10], str(item.thickness),
                    str(item.quantity), f"{item.unit_price:.2f}", f"{item.total_price:.2f}",
                ])
            col_widths = [20, 90, 40, 60, 60, 50, 30, 30, 50, 50]
        else:
            headers = ["№", "Назва", "Тип", "Розміри", "Матеріал", "Товщ.", "К-ть", "Ціна", "Сума"]
            data = [headers]
            for i, item in enumerate(items, 1):
                data.append([
                    str(i), item.name[:30], item.product_type[:15], item.dimensions[:20],
                    item.material[:12], str(item.thickness), str(item.quantity),
                    f"{item.unit_price:.2f}", f"{item.total_price:.2f}",
                ])
            col_widths = [25, 110, 70, 80, 60, 35, 35, 55, 55]
        table = Table(data, colWidths=col_widths, repeatRows=1)
        table_style = TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1565C0")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("ALIGN", (1, 1), (1, -1), "LEFT"),
            ("FONTNAME", (0, 0), (-1, 0), _PDF_FONT_NAME + "-Bold" if _PDF_FONT_NAME != "Helvetica" else "Helvetica-Bold"),
            ("FONTNAME", (0, 1), (-1, -1), _PDF_FONT_NAME),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f5f5")]),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ])
        table.setStyle(table_style)
        story.append(table)
        story.append(Spacer(1, 10))
        total = sum(i.total_price for i in items)
        total_style = ParagraphStyle("TotalStyle", parent=styles["Heading3"], fontName=_PDF_FONT_NAME)
        story.append(Paragraph(f"<b>Всього: {total:,.2f} грн</b>", total_style))
        doc.build(story)

    @staticmethod
    def to_html(items: list[PriceItem], internal: bool = True, title: str = "Прайс-лист") -> str:
        rows = ""
        if internal:
            for i, item in enumerate(items, 1):
                rows += f"""
                <tr>
                    <td>{i}</td><td>{item.name}</td><td>{item.category}</td><td>{item.product_type}</td>
                    <td>{item.dimensions}</td><td>{item.material}</td><td>{item.thickness}</td>
                    <td>{item.unit}</td><td>{item.quantity}</td>
                    <td>{item.cost_price:.2f}</td><td>{item.labor_cost:.2f}</td>
                    <td>{item.overhead_cost:.2f}</td><td>{item.markup_percent:.1f}%</td>
                    <td>{item.unit_price:.2f}</td><td>{item.total_price:.2f}</td>
                    <td>{item.profit:.2f}</td><td>{item.supplier}</td><td>{item.notes_internal}</td>
                </tr>"""
            headers = """
                <th>№</th><th>Назва</th><th>Категорія</th><th>Тип</th><th>Розміри</th>
                <th>Матеріал</th><th>Товщ.</th><th>Од.</th><th>К-ть</th>
                <th>Собіварт.</th><th>Роботи</th><th>Накладні</th><th>Націнка</th>
                <th>Ціна од.</th><th>Сума</th><th>Прибуток</th><th>Постач.</th><th>Примітки</th>
            """
        else:
            for i, item in enumerate(items, 1):
                rows += f"""
                <tr>
                    <td>{i}</td><td>{item.name}</td><td>{item.product_type}</td>
                    <td>{item.dimensions}</td><td>{item.material}</td><td>{item.thickness}</td>
                    <td>{item.unit}</td><td>{item.quantity}</td>
                    <td>{item.unit_price:.2f}</td><td>{item.total_price:.2f}</td>
                    <td>{item.notes_public}</td>
                </tr>"""
            headers = """
                <th>№</th><th>Назва</th><th>Тип</th><th>Розміри</th>
                <th>Матеріал</th><th>Товщ.</th><th>Од.</th><th>К-ть</th>
                <th>Ціна од.</th><th>Сума</th><th>Примітки</th>
            """
        total = sum(i.total_price for i in items)
        total_profit = sum(i.profit for i in items) if internal else 0
        return f"""<!DOCTYPE html>
<html lang="uk">
<head>
    <meta charset="UTF-8">
    <title>{title}</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 20px; }}
        h1 {{ color: #1565C0; text-align: center; }}
        .meta {{ text-align: center; color: #666; margin-bottom: 20px; }}
        table {{ border-collapse: collapse; width: 100%; font-size: 11px; }}
        th {{ background: #1565C0; color: white; padding: 8px; text-align: center; }}
        td {{ border: 1px solid #ccc; padding: 6px; text-align: center; }}
        tr:nth-child(even) {{ background: #f5f5f5; }}
        .total {{ font-weight: bold; font-size: 14px; margin-top: 15px; text-align: right; }}
        .profit {{ color: #2E7D32; }}
        @media print {{ body {{ margin: 0; }} }}
    </style>
</head>
<body>
    <h1>🏷️ {title}</h1>
    <p class="meta">Дата: {datetime.now().strftime('%d.%m.%Y %H:%M')}</p>
    <table>
        <thead><tr>{headers}</tr></thead>
        <tbody>{rows}</tbody>
    </table>
    <p class="total">Всього: <b>{total:,.2f} грн</b></p>
    {f'<p class="total profit">Прибуток: <b>{total_profit:,.2f} грн</b></p>' if internal else ''}
</body>
</html>"""

