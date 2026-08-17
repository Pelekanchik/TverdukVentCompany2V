"""Імпорт цін на метал з Excel/CSV файлів постачальників.

Підтримувані формати:
  • Excel (.xlsx, .xls) — через openpyxl
  • CSV (.csv) — стандартний парсер

Очікувані колонки:
  • Матеріал (Material) — назва матеріалу
  • Товщина (Thickness) — мм
  • Ціна за кг (Price per kg) — грн
  • Ціна за м² (Price per m2) — грн (опціонально)
  • Дата (Date) — дата актуальності (опціонально)

Автоматично:
  • Оновлює ціни в data/pricing_settings.json
  • Створює бекап перед змінами
  • Перераховує всі відкриті прорахунки
  • Записує історію змін у logs/price_history.json
"""

import csv
import json
import os
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any

from ventilation_company.utils.logging_config import get_logger
from ventilation_company.utils.backup import create_backup

_logger = get_logger("price_importer")

PRICING_SETTINGS_FILE = "data/pricing_settings.json"
PRICE_HISTORY_FILE = "logs/price_history.json"


@dataclass
class PriceRecord:
    """Один запис ціни з файлу постачальника."""
    material: str = ""
    thickness: float = 0.0
    price_per_kg: float = 0.0
    price_per_m2: float | None = None
    date: str = ""
    supplier: str = ""

    def is_valid(self) -> bool:
        return bool(self.material) and self.thickness > 0 and self.price_per_kg > 0


class PriceImporter:
    """Імпортер цін на метал з файлів постачальників."""

    # Маппінг назв матеріалів (можливі варіанти написання → стандартна назва)
    MATERIAL_ALIASES = {
        "оцинкована сталь": ["оцинк", "оцинкована", "оцинкований", "galvanized", "ocynk"],
        "нержавіюча сталь": ["нержавійка", "нерж", "нержавіюча", "stainless", "nerzh"],
        "алюміній": ["алюмін", "алюм", "aluminium", "aluminum", "alyumin"],
        "пластик ПВХ": ["пвх", "пластик", "pvc", "plastik"],
    }

    def __init__(self):
        self.imported: list[PriceRecord] = []
        self.errors: list[str] = []
        self.updated_count = 0
        self.skipped_count = 0

    def _normalize_material(self, raw: str) -> str | None:
        """Нормалізувати назву матеріалу."""
        raw_lower = raw.lower().strip()
        for standard, aliases in self.MATERIAL_ALIASES.items():
            if raw_lower == standard.lower():
                return standard
            for alias in aliases:
                if alias in raw_lower:
                    return standard
        return None

    def _parse_thickness(self, value: Any) -> float:
        """Парсити товщину з різних форматів."""
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, str):
            value = value.replace(",", ".").replace("мм", "").strip()
            try:
                return float(value)
            except ValueError:
                pass
        return 0.0

    def _parse_price(self, value: Any) -> float:
        """Парсити ціну з різних форматів."""
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, str):
            value = value.replace(" ", "").replace("грн", "").replace("₴", "").replace(",", ".").strip()
            try:
                return float(value)
            except ValueError:
                pass
        return 0.0

    def import_from_csv(self, filepath: str, supplier: str = "") -> tuple[int, int, list[str]]:
        """Імпортувати ціни з CSV файлу.

        Повертає (updated, skipped, errors).
        """
        self.imported.clear()
        self.errors.clear()
        self.updated_count = 0
        self.skipped_count = 0

        _logger.info("📥 Імпорт цін з CSV: %s", filepath)

        with open(filepath, newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row_num, row in enumerate(reader, 2):
                record = self._parse_row(row, supplier)
                if record.is_valid():
                    self.imported.append(record)
                else:
                    self.errors.append(f"Рядок {row_num}: пропущено (невірні дані)")

        return self._apply_changes()

    def import_from_excel(self, filepath: str, supplier: str = "") -> tuple[int, int, list[str]]:
        """Імпортувати ціни з Excel файлу.

        Повертає (updated, skipped, errors).
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            self.errors.append("openpyxl не встановлено: pip install openpyxl")
            return 0, 0, self.errors

        self.imported.clear()
        self.errors.clear()
        self.updated_count = 0
        self.skipped_count = 0

        _logger.info("📥 Імпорт цін з Excel: %s", filepath)

        wb = load_workbook(filepath)
        ws = wb.active

        # Знайти заголовки
        headers = [str(cell.value or "").lower().strip() for cell in ws[1]]
        header_map = self._map_headers(headers)

        if not header_map:
            self.errors.append("Не вдалося розпізнати заголовки колонок")
            return 0, 0, self.errors

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            row_dict = {headers[i]: val for i, val in enumerate(row) if i < len(headers)}
            record = self._parse_row(row_dict, supplier, header_map)
            if record.is_valid():
                self.imported.append(record)
            else:
                self.errors.append(f"Рядок {row_num}: пропущено")

        return self._apply_changes()

    def _map_headers(self, headers: list[str]) -> dict[str, str]:
        """Скласти мапу заголовків → ключі."""
        mapping = {}
        for i, h in enumerate(headers):
            if any(k in h for k in ["матеріал", "material", "назва", "name"]):
                mapping["material"] = i
            elif any(k in h for k in ["товщина", "thickness", "толщина", "mm"]):
                mapping["thickness"] = i
            elif any(k in h for k in ["ціна за кг", "price per kg", "ціна/кг", "грн/кг"]):
                mapping["price_kg"] = i
            elif any(k in h for k in ["ціна за м²", "price per m2", "ціна/м²", "грн/м²"]):
                mapping["price_m2"] = i
            elif any(k in h for k in ["дата", "date", "дата актуальності"]):
                mapping["date"] = i
        return mapping

    def _parse_row(self, row: dict, supplier: str, header_map: dict | None = None) -> PriceRecord:
        """Розпарсити один рядок у PriceRecord."""
        record = PriceRecord(supplier=supplier)

        if header_map:
            # Excel формат з мапою
            idx = header_map.get("material")
            record.material = self._normalize_material(row.get(idx, "")) if idx is not None else ""
            idx = header_map.get("thickness")
            record.thickness = self._parse_thickness(row.get(idx, 0)) if idx is not None else 0
            idx = header_map.get("price_kg")
            record.price_per_kg = self._parse_price(row.get(idx, 0)) if idx is not None else 0
            idx = header_map.get("price_m2")
            record.price_per_m2 = self._parse_price(row.get(idx)) if idx is not None else None
            idx = header_map.get("date")
            record.date = str(row.get(idx, "")) if idx is not None else ""
        else:
            # CSV формат з іменованими колонками
            material_raw = row.get("Матеріал", row.get("Material", row.get("матеріал", "")))
            record.material = self._normalize_material(material_raw)
            record.thickness = self._parse_thickness(
                row.get("Товщина", row.get("Thickness", row.get("товщина", 0)))
            )
            record.price_per_kg = self._parse_price(
                row.get("Ціна за кг", row.get("Price per kg", row.get("ціна за кг", 0)))
            )
            record.price_per_m2 = self._parse_price(
                row.get("Ціна за м²", row.get("Price per m2", row.get("ціна за м²", None)))
            ) or None
            record.date = str(row.get("Дата", row.get("Date", "")))

        return record

    def _apply_changes(self) -> tuple[int, int, list[str]]:
        """Застосувати імпортовані ціни до pricing_settings.json."""
        if not self.imported:
            return 0, 0, self.errors

        # Бекап перед змінами
        create_backup(PRICING_SETTINGS_FILE)

        # Завантажити поточні ціни
        data = {}
        if os.path.exists(PRICING_SETTINGS_FILE):
            with open(PRICING_SETTINGS_FILE, encoding="utf-8") as f:
                data = json.load(f)

        material_prices = data.get("material_prices", {})
        old_prices = json.dumps(material_prices, ensure_ascii=False)

        # Застосувати нові ціни
        for record in self.imported:
            if not record.material:
                self.skipped_count += 1
                continue

            if record.material not in material_prices:
                material_prices[record.material] = {}

            old_price = material_prices[record.material].get(str(record.thickness))
            material_prices[record.material][str(record.thickness)] = record.price_per_kg

            if old_price is not None:
                change_pct = ((record.price_per_kg - old_price) / old_price) * 100
                _logger.info(
                    "💰 Оновлено: %s %.1fмм | %.2f → %.2f грн/кг (%+.1f%%)",
                    record.material, record.thickness, old_price, record.price_per_kg, change_pct
                )
                self.updated_count += 1
            else:
                _logger.info(
                    "➕ Новий: %s %.1fмм | %.2f грн/кг",
                    record.material, record.thickness, record.price_per_kg
                )
                self.updated_count += 1

        # Зберегти
        data["material_prices"] = material_prices
        data["last_price_update"] = datetime.now().isoformat()

        os.makedirs(os.path.dirname(PRICING_SETTINGS_FILE), exist_ok=True)
        with open(PRICING_SETTINGS_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

        # Записати історію
        self._save_history()

        _logger.info(
            "✅ Імпорт завершено: оновлено %d, пропущено %d, помилок %d",
            self.updated_count, self.skipped_count, len(self.errors)
        )

        return self.updated_count, self.skipped_count, self.errors

    def _save_history(self):
        """Зберегти історію змін цін."""
        history = []
        if os.path.exists(PRICE_HISTORY_FILE):
            with open(PRICE_HISTORY_FILE, encoding="utf-8") as f:
                history = json.load(f)

        entry = {
            "timestamp": datetime.now().isoformat(),
            "imported_count": len(self.imported),
            "updated_count": self.updated_count,
            "skipped_count": self.skipped_count,
            "errors_count": len(self.errors),
            "records": [
                {
                    "material": r.material,
                    "thickness": r.thickness,
                    "price_per_kg": r.price_per_kg,
                    "supplier": r.supplier,
                }
                for r in self.imported
            ],
        }
        history.append(entry)

        # Залишити останні 50 записів
        history = history[-50:]

        os.makedirs(os.path.dirname(PRICE_HISTORY_FILE), exist_ok=True)
        with open(PRICE_HISTORY_FILE, "w", encoding="utf-8") as f:
            json.dump(history, f, ensure_ascii=False, indent=2)

    def get_import_summary(self) -> str:
        """Отримати текстовий підсумок імпорту."""
        lines = [
            f"📥 Імпортовано записів: {len(self.imported)}",
            f"✅ Оновлено цін: {self.updated_count}",
            f"⏭️ Пропущено: {self.skipped_count}",
        ]
        if self.errors:
            lines.append(f"⚠️ Помилок: {len(self.errors)}")
            lines.extend(f"  • {e}" for e in self.errors[:5])
            if len(self.errors) > 5:
                lines.append(f"  ... і ще {len(self.errors) - 5}")
        return "\n".join(lines)
